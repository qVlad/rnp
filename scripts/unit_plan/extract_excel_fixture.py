"""Extract contract-test fixture from LeymanKids UNIT-план Excel.

Цель: вытащить N случайных строк из эталонного Excel и сохранить как
JSON-фикстуру для contract-теста `test_compute_row_excel_contract.py`.

Каждая строка содержит:
- input: значения колонок которые подаются на вход `compute_row`
  (volume_l, base_price, discount_pct, СПП, литры, склад и т.д.)
- expected: ожидаемые значения output-полей `UnitPlanRowDTO`
  (price_after_*, commission_rub, logistics_rub, ..., profit_rub, margin_pct)
- global_constants: значения из R1 Excel (на момент снапшота)

Usage:
    python scripts/unit_plan/extract_excel_fixture.py \\
        /Users/user/Downloads/LeymanKids\\ UNIT_план\\ WB\\ Обновление.xlsx \\
        --output backend/tests/fixtures/unit_plan_excel_50.json \\
        --rows 50 \\
        --seed 42

Файл `xlsx` обрабатывается через openpyxl с `data_only=True` (берём
вычисленные значения формул, не сами формулы).
"""
from __future__ import annotations

import argparse
import json
import random
import sys
import warnings
from decimal import Decimal
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore", message="Unknown extension is not supported")


# Excel column letter → DTO field name. Из UNIT_PLAN.md §4.
# Только те колонки которые нам реально нужны для contract-теста.
INPUT_COLUMNS = {
    "A": "warehouse",
    "B": "volume_l",
    "C": "brand",
    "D": "subject",
    "E": "vendor_code",
    "F": "nm_id",
    "G": "stock_wb",
    "H": "stock_fbs",
    "K": "base_price",
    "L": "discount_pct",   # 0-1
    "R": "spp_pct",        # 0-1
    "Y": "is_fbs_str",     # "Да"/"Нет"
    "AA": "is_monopallet_str",
    "AB": "items_per_monopallet",
    "AD": "buyout_pct",    # 0-1
    "AE": "warehouse_coef_pct",
    "AK": "cogs_rub",
}

EXPECTED_COLUMNS = {
    "I": "stock_effective",
    "J": "days_to_stockout",
    "O": "price_after_discount",
    "Q": "price_after_wb_club",
    "S": "price_after_spp",
    "T": "price_final",
    "U": "commission_pct",          # 0-1
    "W": "commission_total_pct",
    "X": "commission_rub",
    "Z": "logistics_box_rub",
    "AC": "logistics_pallet_rub",
    "AF": "logistics_rub",
    "AG": "reverse_logistics_rub",
    "AH": "logistics_share",
    "AI": "storage_rub",
    "AJ": "storage_share",
    "AL": "cogs_share",
    "AM": "marketing_rub",
    "AO": "tax_rub",
    "AQ": "vat_rub",
    "AS": "acceptance_rub",
    "AU": "profit_rub",
    "AV": "margin_pct",
    "AW": "roi_pct",
}

# Глобальные константы из R1 (см. UNIT_PLAN.md §2).
GLOBAL_CONSTANTS_CELLS = {
    "J1": "velocity_days",
    "P1": "wb_club_pct",
    "R1": "spp_default_pct",
    "T1": "wb_wallet_pct",
    "V1": "acquiring_pct",
    "AA1": "il_coef",
    "AC1": "irp_coef",
    "AN1": "marketing_pct",
    "AP1": "tax_pct",
    "AQ1": "vat_mode_str",   # "Да, включаем" / "Да, не включаем" / "не включаем"
    "AR1": "vat_pct",
    "AS1": "acceptance_rub_per_liter",
}


def _serialize(value):
    """Convert openpyxl Cell value to JSON-safe representation."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        # Числа сохраняем как строки чтобы избежать float-imprecision
        # в JSON. Decimal на той стороне реконструируется через Decimal(str(x)).
        return str(value)
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, bool):
        return value
    return str(value)


def _load_box_tariffs(wb) -> dict[str, dict]:
    """Map warehouse_name → box tariff inputs from `Логистика и хранение короб` sheet.

    Columns: A=warehouse, B=коэф_логистики_%, C=Логистика_за_1л,
             D=Доп_л, E=коэф_хранения_%, F=Хранение_за_1л, G=Доп_л.
    """
    ws = wb["Логистика и хранение короб"]
    out: dict[str, dict] = {}
    for r in range(2, ws.max_row + 1):
        wh = ws.cell(r, 1).value
        if not wh:
            continue
        out[str(wh)] = {
            "delivery_expr_pct": _serialize(ws.cell(r, 2).value),  # %, например 160 → 1.60
            "delivery_base": _serialize(ws.cell(r, 3).value),
            "delivery_liter": _serialize(ws.cell(r, 4).value),
            "storage_expr_pct": _serialize(ws.cell(r, 5).value),
            "storage_base": _serialize(ws.cell(r, 6).value),
            "storage_liter": _serialize(ws.cell(r, 7).value),
        }
    return out


def extract(excel_path: Path, n_rows: int, seed: int) -> dict:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["UNIT"]
    box_tariffs = _load_box_tariffs(wb)

    # 1. Глобальные константы
    constants = {}
    for cell, key in GLOBAL_CONSTANTS_CELLS.items():
        constants[key] = _serialize(ws[cell].value)

    # 2. Все строки с непустым nm_id (колонка F)
    all_rows = []
    for r in range(3, ws.max_row + 1):
        nm = ws.cell(r, 6).value  # F
        if nm is None:
            continue
        # Skip rows с пустыми ключевыми полями (volume / base_price / cogs)
        if ws.cell(r, 2).value is None or ws.cell(r, 11).value is None:
            continue
        if ws.cell(r, 37).value is None:  # AK = cogs
            continue
        all_rows.append(r)

    print(f"Total candidate rows (nm_id + volume + price + cogs all present): {len(all_rows)}")
    if len(all_rows) < n_rows:
        print(f"⚠ Only {len(all_rows)} candidates available; reducing n_rows.")
        n_rows = len(all_rows)

    # 3. Случайная выборка
    rng = random.Random(seed)
    sample_row_nums = rng.sample(all_rows, n_rows)
    sample_row_nums.sort()

    rows = []
    for r in sample_row_nums:
        input_data = {}
        for col_letter, field in INPUT_COLUMNS.items():
            cell = ws[f"{col_letter}{r}"]
            input_data[field] = _serialize(cell.value)
        expected_data = {}
        for col_letter, field in EXPECTED_COLUMNS.items():
            cell = ws[f"{col_letter}{r}"]
            expected_data[field] = _serialize(cell.value)
        warehouse = str(ws.cell(r, 1).value or "")
        rows.append({
            "row_num": r,
            "nm_id": _serialize(ws.cell(r, 6).value),
            "input": input_data,
            "box_tariff": box_tariffs.get(warehouse, {}),
            "expected": expected_data,
        })

    return {
        "source": "LeymanKids UNIT_план WB Обновление.xlsx",
        "extracted_at": "2026-05-19",
        "seed": seed,
        "n_rows": len(rows),
        "global_constants": constants,
        "rows": rows,
    }


def main(argv: list[str]) -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("excel", type=Path, help="Path to LeymanKids xlsx")
    p.add_argument("--output", type=Path, required=True)
    p.add_argument("--rows", type=int, default=50)
    p.add_argument("--seed", type=int, default=42)
    args = p.parse_args(argv)

    if not args.excel.exists():
        print(f"❌ Excel not found: {args.excel}", file=sys.stderr)
        return 1

    fixture = extract(args.excel, args.rows, args.seed)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(fixture, ensure_ascii=False, indent=2))
    print(f"✓ Wrote {len(fixture['rows'])} rows to {args.output}")
    print(f"  Global constants: {fixture['global_constants']}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
