"""Unit-тесты `services.unit_plan_xlsx.build_unit_plan_xlsx` (UNIT-PLAN-014).

Цель: проверить, что XLSX-export 1:1 совпадает по структуре с эталонным
LeymanKids UNIT_план — единственный лист «UNIT», R1 — константы в нужных
ячейках, R2 — headers, R3+ — данные.

Тесты — чисто-функциональные: `build_unit_plan_xlsx` принимает list[dict]
и возвращает bytes; БД/сессия не нужны.
"""
from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from app.services.unit_plan_xlsx import (
    COLUMNS,
    SHEET_NAME,
    build_unit_plan_xlsx,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _global_cfg_dict() -> dict[str, object]:
    """Минимальный валидный конфиг для R1."""
    return {
        "velocity_days": 30,
        "wb_club_pct": Decimal("0.0"),
        "spp_default_pct": Decimal("0.2"),
        "wb_wallet_pct": Decimal("0.02"),
        "acquiring_pct": Decimal("0.02"),
        "il_coef": Decimal("1.16"),
        "irp_coef": Decimal("0.017"),
        "marketing_pct": Decimal("0.03"),
        "tax_pct": Decimal("0.08"),
        "vat_mode": "none",
        "vat_pct": Decimal("0.1"),
        "acceptance_rub_per_liter": Decimal("1.7"),
        "acceptance_multiplier": Decimal("1.0"),
        "buyout_fallback_pct": Decimal("0.5"),
        "storage_days": 1,
    }


def _fake_row(
    *,
    nm_id: int,
    vendor_code: str,
    brand: str = "BrandA",
    subject: str = "Платье",
    warehouse: str = "Электросталь",
) -> dict[str, object]:
    """Сериализованный `UnitPlanRowDTO` (все 51 поле).

    Числа храним как Decimal/int — `build_unit_plan_xlsx` приводит их к float
    с нужным number_format. `is_fbs`/`is_monopallet` — bool.
    """
    return {
        "nm_id": nm_id,
        "vendor_code": vendor_code,
        "brand": brand,
        "subject": subject,
        "warehouse": warehouse,
        "volume_l": Decimal("0.5"),
        "stock_wb": 100,
        "stock_fbs": 0,
        "stock_effective": Decimal("150.0"),
        "days_to_stockout": Decimal("45.0"),
        "base_price": Decimal("2000.0"),
        "discount_pct": Decimal("0.3"),
        "price_after_discount": Decimal("1400.0"),
        "wb_club_pct": Decimal("0.0"),
        "price_after_wb_club": Decimal("1400.0"),
        "spp_pct": Decimal("0.28"),
        "price_after_spp": Decimal("1008.0"),
        "wb_wallet_pct": Decimal("0.02"),
        "price_final": Decimal("987.84"),
        "commission_pct": Decimal("0.18"),
        "acquiring_pct": Decimal("0.02"),
        "commission_total_pct": Decimal("0.20"),
        "commission_rub": Decimal("280.0"),
        "is_fbs": False,
        "is_monopallet": False,
        "items_per_monopallet": None,
        "buyout_pct": Decimal("0.5"),
        "warehouse_coef_pct": Decimal("1.5"),
        "logistics_box_rub": Decimal("47.4"),
        "logistics_pallet_rub": Decimal("0"),
        "reverse_logistics_rub": Decimal("29.0"),
        "logistics_rub": Decimal("75.8"),
        "logistics_share": Decimal("0.054"),
        "storage_rub": Decimal("12.0"),
        "storage_share": Decimal("0.0086"),
        "cogs_rub": Decimal("400.0"),
        "cogs_share": Decimal("0.286"),
        "marketing_rub": Decimal("42.0"),
        "marketing_pct": Decimal("0.03"),
        "tax_rub": Decimal("79.0"),
        "tax_pct": Decimal("0.08"),
        "vat_rub": Decimal("0"),
        "vat_pct": Decimal("0.1"),
        "acceptance_rub": Decimal("1.7"),
        "acceptance_share": Decimal("0.0012"),
        "profit_rub": Decimal("510.1"),
        "margin_pct": Decimal("0.364"),
        "roi_pct": Decimal("1.275"),
        "abc_label": "A",
        "season_label": "о-в",
        "gender_label": "д",
    }


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_xlsx_structure_one_sheet_named_unit() -> None:
    """Файл содержит ровно один лист с именем «UNIT»."""
    rows = [
        _fake_row(nm_id=111, vendor_code="A-1"),
    ]
    data = build_unit_plan_xlsx(rows=rows, config=_global_cfg_dict())
    assert isinstance(data, bytes) and len(data) > 100

    wb = load_workbook(BytesIO(data))
    assert wb.sheetnames == [SHEET_NAME]


def test_r1_global_constants_in_fixed_cells() -> None:
    """R1 — константы в тех же ячейках, что в эталоне LeymanKids."""
    config = _global_cfg_dict()
    rows = [_fake_row(nm_id=111, vendor_code="A-1")]
    wb = load_workbook(BytesIO(build_unit_plan_xlsx(rows=rows, config=config)))
    ws = wb[SHEET_NAME]

    # E1 — лейбл, F1 — сегодняшняя дата.
    assert isinstance(ws["E1"].value, str) and "Обновлен" in ws["E1"].value
    # F1 может быть date или datetime — главное, не пусто.
    assert ws["F1"].value is not None

    # J1 = velocity_days (30).
    assert ws["J1"].value == 30
    # P1 = wb_club_pct (0.0).
    assert ws["P1"].value == 0.0
    # R1 = spp_default_pct (0.2).
    assert ws["R1"].value == 0.2
    # T1 = wb_wallet_pct (0.02).
    assert ws["T1"].value == 0.02
    # V1 = acquiring_pct (0.02).
    assert ws["V1"].value == 0.02
    # Z1 = «ИЛ», AA1 = il_coef (1.16).
    assert ws["Z1"].value == "ИЛ"
    assert ws["AA1"].value == 1.16
    # AB1 = «ИРП», AC1 = irp_coef (0.017).
    assert ws["AB1"].value == "ИРП"
    assert abs(float(ws["AC1"].value) - 0.017) < 1e-9
    # AN1 = marketing_pct (0.03).
    assert ws["AN1"].value == 0.03
    # AP1 = tax_pct (0.08).
    assert ws["AP1"].value == 0.08
    # AQ1 = «не включаем» (vat_mode='none').
    assert ws["AQ1"].value == "не включаем"
    # AR1 = vat_pct (0.1).
    assert ws["AR1"].value == 0.1
    # AS1 = acceptance_rub_per_liter (1.7).
    assert abs(float(ws["AS1"].value) - 1.7) < 1e-9


def test_r2_headers_match_reference() -> None:
    """R2 — точные русские названия колонок (sample-check)."""
    rows = [_fake_row(nm_id=111, vendor_code="A-1")]
    wb = load_workbook(BytesIO(build_unit_plan_xlsx(rows=rows, config=_global_cfg_dict())))
    ws = wb[SHEET_NAME]

    assert ws["A2"].value == "Склад"
    assert ws["B2"].value == "Литры"
    assert ws["C2"].value == "Бренд"
    assert ws["D2"].value == "Предмет"
    assert ws["F2"].value == "Артикул WB"
    assert ws["G2"].value == "Остаток на всех складах WB"
    assert ws["J2"].value == "закончится через, дн"
    assert ws["O2"].value == "Цена продажи (без СПП)"
    assert ws["P2"].value == "ВБ Клуб"
    assert ws["R2"].value == "Размер СПП"
    assert ws["T2"].value == "Цена с WB кошелек"
    assert ws["Y2"].value == "FBS"
    assert ws["AA2"].value == "Монопаллеты (Да/Нет)"
    assert ws["AD2"].value == "% выкупа"
    assert ws["AU2"].value == "Прибыль в рублях"
    assert ws["AV2"].value == "Маржинальность, прибыль % от продажи"
    assert ws["AW2"].value == "Рентабельность, прибыль % от сс"
    assert ws["AZ2"].value == "АВС"
    # последняя BF — header без даты (она задаётся query-param forecast_date)
    assert ws["BF2"].value == "Прогноз остатка"


def test_data_rows_count_and_values() -> None:
    """3 fake-rows → 3 строки данных, начиная с R3. Spot-check значений."""
    rows = [
        _fake_row(nm_id=11111, vendor_code="V-1", brand="BrandA"),
        _fake_row(nm_id=22222, vendor_code="V-2", brand="BrandB"),
        _fake_row(nm_id=33333, vendor_code="V-3", brand="BrandC", warehouse="Коледино"),
    ]
    wb = load_workbook(BytesIO(build_unit_plan_xlsx(rows=rows, config=_global_cfg_dict())))
    ws = wb[SHEET_NAME]

    # max_row == 2 (constants + headers) + len(rows) = 5.
    assert ws.max_row == 2 + len(rows)

    # R3.F = nm_id первой строки.
    assert ws["F3"].value == 11111
    assert ws["F4"].value == 22222
    assert ws["F5"].value == 33333

    # R3.A = warehouse, R3.C = brand.
    assert ws["A3"].value == "Электросталь"
    assert ws["A5"].value == "Коледино"
    assert ws["C4"].value == "BrandB"
    assert ws["E3"].value == "V-1"

    # Y3 (is_fbs=False) → "Нет"; AA3 (is_monopallet=False) → "Нет".
    assert ws["Y3"].value == "Нет"
    assert ws["AA3"].value == "Нет"

    # AU3 (profit_rub) — число, формат с 2 знаками.
    assert abs(float(ws["AU3"].value) - 510.1) < 1e-6
    assert ws["AU3"].number_format == "0.00"

    # AV3 (margin_pct) — формат проценты.
    assert ws["AV3"].number_format == "0.00%"
    assert abs(float(ws["AV3"].value) - 0.364) < 1e-6

    # L3 (discount_pct) → 0.3 как доля 0-1, формат '0.00%'.
    assert ws["L3"].value == 0.3
    assert ws["L3"].number_format == "0.00%"

    # N3 — «Проверка». При совпадении L и M → "OK".
    assert ws["N3"].value == "OK"


def test_empty_rows_still_produces_valid_workbook() -> None:
    """Пустой список SKU → файл валидный, R1+R2 на месте, R3 пустая."""
    wb = load_workbook(
        BytesIO(build_unit_plan_xlsx(rows=[], config=_global_cfg_dict()))
    )
    ws = wb[SHEET_NAME]
    assert ws["A2"].value == "Склад"
    # Headers + R1, без данных.
    assert ws.max_row == 2
    # Все колонки описаны:
    assert len(COLUMNS) == 58  # A..AZ (52) + BA..BF (6) = 58 колонок
