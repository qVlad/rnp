"""Тесты парсеров WMS «Свой склад» (TASK-DEV-098). Без БД — чистые функции.

Синтетическая фикстура воспроизводит все особенности реального
`PackingList.xlsx`, найденные при разборе:
  - шапка НЕ в первой строке (в реальном файле — 6-я), двуязычная (`Barcode\\n条码`);
  - `Box Code` продублирован в строках-продолжениях, `No`/вес/габариты — только
    в первой строке короба;
  - **границы физического короба идут по `No`, а не по `Box Code`**: два подряд
    идущих короба могут иметь одинаковый код `—` (в реальном файле — 6 таких);
  - последняя строка `ИТОГО / 合计` должна отбрасываться.

Дополнительно проверяется формат B (колонки «Склад» / «Код ячейки») и
round-trip «экспорт состояния → повторный разбор».
"""
from __future__ import annotations

import io
import os
from pathlib import Path

import pytest
from openpyxl import Workbook

from app.services.warehouse.cells import (
    compute_sort_orders,
    generate_cells,
    parse_cell_code,
    parse_cells_file,
)
from app.services.warehouse.excel import build_state_xlsx
from app.services.warehouse.packing_list import parse_receive_file

# Реальный файл поставщика лежит вне репозитория — тест на нём пропускается,
# если файла нет (CI / чужая машина).
REAL_PACKING_LIST = Path(
    "/Users/user/ai-work/altecom/04_Поставки/2_fixed/output/PackingList.xlsx"
)


def _xlsx(rows: list[list], title: str = "Packing List") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _packing_list_like() -> bytes:
    """Фикстура «как настоящий PackingList»: шапка в 6-й строке + ИТОГО."""
    return _xlsx(
        [
            ["PACKING LIST / 装箱单"],
            ["Сводный по всей поставке (5 коробов)"],
            [],
            [],
            [],
            # шапка — 6-я строка, двуязычная, с мусорными колонками
            [
                "No\n序号",
                "Box Code",
                "Barcode\n条码",
                "Size\n尺码",
                "Qty\n数量",
                "PCS/CTN\n每箱件数",
                "CTN\n箱数",
                "G.W.(kg)\n毛重",
                "CBM\n体积",
            ],
            # короб 1 — сборный, 3 позиции, код только в 1-й строке продублирован
            [1, "ALT-001", "2049332423529", "122", 40, 140, 1, 25.6, 0.096],
            [None, "ALT-001", "2049332423536", "134", 40, None, None, None, None],
            [None, "ALT-001", "2049332423543", "140", 60, None, None, None, None],
            # короб 2 — моно
            [2, "ALT-002", "2049332521584", "110", 20, 20, 2, 12.0, 0.096],
            # коробы 3 и 4 — БЕЗ кода, подряд: разделяются только по `No`
            [3, "—", "2043915143231", "37", 10, 10, 3, 5.0, 0.096],
            [4, "—", "2043915143262", "40", 10, 10, 4, 5.0, 0.096],
            # короб 5 — один баркод дважды (должен просуммироваться → моно)
            [5, "ALT-005", "2049332542275", "146", 30, 60, 5, 24.0, 0.096],
            [None, "ALT-005", "2049332542275", "146", 30, None, None, None, None],
            # итоговая строка — отбрасывается
            [None, "ИТОГО / 合计", None, None, 240, None, 5, None, None],
        ]
    )


def test_header_not_in_first_row_and_totals_dropped() -> None:
    result = parse_receive_file(_packing_list_like(), supply_ref="PL-1")
    stats = result["stats"]
    assert stats["boxes_total"] == 5, "ИТОГО не должно попадать в коробы"
    assert stats["total_qty"] == 240
    assert stats["rows_dropped"] == 0


def test_box_boundaries_follow_no_column_not_box_code() -> None:
    """Два подряд идущих короба с кодом `—` не должны слипнуться в один."""
    result = parse_receive_file(_packing_list_like(), supply_ref="PL-1")
    nocode = [b for b in result["boxes"] if b["box_code_synthetic"]]
    assert len(nocode) == 2
    assert {b["src_no"] for b in nocode} == {3, 4}
    # синтетические коды уникальны и содержат supply_ref
    codes = {b["box_code"] for b in nocode}
    assert codes == {"NOCODE-PL-1-3", "NOCODE-PL-1-4"}
    assert result["stats"]["boxes_without_code"] == 2
    assert result["warnings"], "по коробам без ШК должно быть предупреждение"


def test_continuation_rows_and_mono_detection() -> None:
    result = parse_receive_file(_packing_list_like())
    by_code = {b["box_code"]: b for b in result["boxes"]}

    mixed = by_code["ALT-001"]
    assert len(mixed["items"]) == 3
    assert mixed["is_mono"] is False
    assert mixed["total_qty"] == 140
    # вес/объём берутся из первой строки короба
    assert mixed["gross_weight_kg"] == "25.6"
    assert mixed["cbm"] == "0.096"

    mono = by_code["ALT-002"]
    assert mono["is_mono"] is True
    assert mono["total_qty"] == 20

    # один баркод в двух строках → одна позиция с суммой, короб моно
    dup = by_code["ALT-005"]
    assert len(dup["items"]) == 1
    assert dup["items"][0]["qty"] == 60
    assert dup["is_mono"] is True

    assert result["stats"]["boxes_mono"] == 4  # ALT-002, 2 NOCODE, ALT-005
    assert result["stats"]["boxes_mixed"] == 1


def test_format_b_warehouse_and_cell_columns() -> None:
    """Формат B: «Склад» + «Код ячейки» → адрес короба и пустые ячейки."""
    content = _xlsx(
        [
            ["Склад", "Код ячейки", "No", "Box Code", "Barcode", "Size", "Qty"],
            ["Основной", "A-01-02-01", 1, "ALT-001", "2049332423529", "122", 40],
            ["Основной", None, None, "ALT-001", "2049332423536", "134", 10],
            # без ячейки → на хранение
            ["Основной", None, 2, "ALT-002", "2049332521584", "110", 20],
            # только адрес → объявление пустой ячейки
            ["Основной", "A-01-02-05", None, None, None, None, None],
            ["Тула", "B-01-01-01", 3, "ALT-003", "2043915143231", "37", 5],
        ],
        title="Размещение",
    )
    result = parse_receive_file(content, supply_ref="PL-2")
    by_code = {b["box_code"]: b for b in result["boxes"]}

    assert by_code["ALT-001"]["cell_code"] == "A-01-02-01"
    assert by_code["ALT-001"]["warehouse"] == "Основной"
    assert by_code["ALT-001"]["total_qty"] == 50
    # forward-fill склада на строку-продолжение не сломал границы коробов
    assert by_code["ALT-002"]["cell_code"] is None
    assert by_code["ALT-002"]["warehouse"] == "Основной"
    assert by_code["ALT-003"]["warehouse"] == "Тула"

    assert result["empty_cells"] == [{"cell_code": "A-01-02-05", "warehouse": "Основной"}]
    assert result["stats"]["empty_cells"] == 1


def test_rows_without_barcode_or_qty_are_dropped() -> None:
    content = _xlsx(
        [
            ["No", "Box Code", "Barcode", "Size", "Qty"],
            [1, "ALT-001", "2049332423529", "122", 40],
            [None, "ALT-001", None, "134", 10],   # нет баркода
            [None, "ALT-001", "2049332423543", "140", 0],  # qty = 0
            [None, "ALT-001", "2049332423550", "146", -5],  # отрицательное
        ]
    )
    result = parse_receive_file(content)
    assert result["stats"]["boxes_total"] == 1
    assert result["stats"]["total_qty"] == 40
    assert result["stats"]["rows_dropped"] == 3


def test_barcode_float_tail_is_normalized() -> None:
    """Excel часто отдаёт баркод как float `2049332423529.0`."""
    content = _xlsx(
        [
            ["No", "Box Code", "Barcode", "Qty"],
            [1, "ALT-001", 2049332423529.0, 10],
        ]
    )
    result = parse_receive_file(content)
    assert result["boxes"][0]["items"][0]["barcode"] == "2049332423529"


def test_state_export_import_round_trip() -> None:
    """Экспорт состояния пишет формат B → парсер читает его без потерь."""
    source = parse_receive_file(_packing_list_like(), supply_ref="PL-1")
    rows = [
        {
            "warehouse_name": "Основной",
            "cell_code": "A-01-01-01" if i == 0 else None,
            "src_no": box["src_no"],
            "box_code": box["box_code"],
            "barcode": item["barcode"],
            "size": item["size"],
            "qty": item["qty"],
            "status_label": "На хранении",
            "supply_ref": "PL-1",
        }
        for i, box in enumerate(source["boxes"])
        for item in box["items"]
    ]
    again = parse_receive_file(build_state_xlsx(rows), supply_ref="PL-1")

    for key in ("boxes_total", "boxes_mono", "boxes_mixed", "barcodes_unique", "total_qty"):
        assert again["stats"][key] == source["stats"][key], f"round-trip разошёлся по {key}"
    # адрес и склад сохранились
    placed = [b for b in again["boxes"] if b["cell_code"]]
    assert len(placed) == 1
    assert placed[0]["cell_code"] == "A-01-01-01"
    assert all(b["warehouse"] == "Основной" for b in again["boxes"])


# --------------------------------------------------------------------------
# Сетка ячеек (формат A) + порядок обхода
# --------------------------------------------------------------------------


def test_parse_cell_code_variants() -> None:
    assert parse_cell_code("A-01-02-03") == {
        "zone": "A",
        "rack": "01",
        "level": "02",
        "pos": "03",
    }
    assert parse_cell_code("B/3/1/5") == {"zone": "B", "rack": "3", "level": "1", "pos": "5"}
    assert parse_cell_code("ZONA-1") == {
        "zone": "ZONA",
        "rack": "1",
        "level": None,
        "pos": None,
    }


def test_sort_order_is_numeric_not_lexicographic() -> None:
    """`A-02` должен идти раньше `A-10` — иначе маршрут отбора будет петлять."""
    cells = [
        {"code": "A-10-01-01", "zone": "A", "rack": "10", "level": "01", "pos": "01"},
        {"code": "A-02-01-01", "zone": "A", "rack": "02", "level": "01", "pos": "01"},
    ]
    compute_sort_orders(cells)
    by_code = {c["code"]: c["sort_order"] for c in cells}
    assert by_code["A-02-01-01"] < by_code["A-10-01-01"]


def test_generate_cells_grid() -> None:
    cells = generate_cells("A", racks=2, levels=2, positions=3)
    assert len(cells) == 12
    assert cells[0]["code"] == "A-01-01-01"
    assert cells[-1]["code"] == "A-02-02-03"
    # sort_order строго возрастает по маршруту
    orders = [c["sort_order"] for c in sorted(cells, key=lambda c: c["sort_order"])]
    assert orders == sorted(orders)
    assert len(set(orders)) == len(orders)


def test_generate_cells_rejects_absurd_grid() -> None:
    with pytest.raises(ValueError):
        generate_cells("A", racks=0, levels=1, positions=1)
    with pytest.raises(ValueError):
        generate_cells("A", racks=100, levels=100, positions=100)


def test_parse_cells_file_format_a() -> None:
    content = _xlsx(
        [
            ["Склад", "Код ячейки", "Зона", "Активна", "Примечание"],
            ["Основной", "A-01-02-01", "A", "да", None],
            ["Основной", "A-01-02-02", None, None, "зона парсится из кода"],
            [None, "A-01-02-03", None, "нет", None],  # склад forward-fill
            ["Тула", "B-03-01-05", "B", "да", "верхний ярус"],
            ["Основной", "A-01-02-01", "A", "да", None],  # дубль → пропуск
            ["Основной", None, None, None, None],  # без кода → пропуск
        ],
        title="Ячейки",
    )
    result = parse_cells_file(content)
    codes = [c["code"] for c in result["cells"]]
    assert codes.count("A-01-02-01") == 1
    assert result["stats"]["cells_total"] == 4
    assert result["stats"]["duplicates"] == 1
    assert result["stats"]["rows_dropped"] == 1
    assert result["stats"]["warehouses"] == ["Основной", "Тула"]

    by_code = {c["code"]: c for c in result["cells"]}
    assert by_code["A-01-02-02"]["zone"] == "A"  # из кода
    assert by_code["A-01-02-03"]["warehouse"] == "Основной"  # forward-fill
    assert by_code["A-01-02-03"]["is_active"] is False


# --------------------------------------------------------------------------
# Реальный файл поставщика — эталонные цифры
# --------------------------------------------------------------------------


@pytest.mark.skipif(
    not REAL_PACKING_LIST.exists(),
    reason="реальный PackingList.xlsx недоступен (лежит вне репозитория)",
)
def test_real_packing_list_reference_numbers() -> None:
    """Эталон на живом файле поставщика (822 короба / 451 баркод / 56 983 шт)."""
    stats = parse_receive_file(REAL_PACKING_LIST.read_bytes(), supply_ref="PL")["stats"]
    assert stats["boxes_total"] == 822
    assert stats["boxes_mono"] == 565
    assert stats["boxes_mixed"] == 257
    assert stats["boxes_without_code"] == 6
    assert stats["barcodes_unique"] == 451
    assert stats["total_qty"] == 56_983  # сходится со строкой ИТОГО
    assert stats["rows_dropped"] == 0


@pytest.mark.skipif(
    not REAL_PACKING_LIST.exists(),
    reason="реальный PackingList.xlsx недоступен (лежит вне репозитория)",
)
def test_real_packing_list_mono_coverage_gap() -> None:
    """Моно-короба покрывают лишь часть ассортимента.

    Обоснование алгоритма размещения: 565 моно-коробов дают всего 287 разных
    баркодов, а 164 баркода существуют ТОЛЬКО в сборных коробах. Поэтому
    greedy-шаг по сборным коробам обязателен, иначе треть ассортимента никогда
    не попадёт в зону отбора.
    """
    boxes = parse_receive_file(REAL_PACKING_LIST.read_bytes(), supply_ref="PL")["boxes"]
    all_barcodes = {i["barcode"] for b in boxes for i in b["items"]}
    mono_barcodes = {b["items"][0]["barcode"] for b in boxes if b["is_mono"]}
    assert len(all_barcodes) == 451
    assert len(mono_barcodes) == 287
    assert len(all_barcodes - mono_barcodes) == 164


# --------------------------------------------------------------------------
# Второй формат: лист «короба» своего склада
# --------------------------------------------------------------------------

REAL_STORAGE_FILE = Path("/Users/user/Downloads/Хранение собственный склад ИНК.xlsx")


def _boxes_sheet_like() -> bytes:
    """Формат «короба»: `Дата упаковки | Номер короба | Баркод | Количество | Наименование | Размер`.

    Отличия от PackingList: другие названия колонок, номер короба числовой и
    стоит только в первой строке короба (в PackingList он продублирован).
    """
    return _xlsx(
        [
            [
                "Дата упаковки",
                "Номер короба",
                "Баркод",
                "Количество",
                None,
                "Наименование",
                "Размер",
            ],
            ["09.04", 7, 2038228335909, 65, None, "Пижама_комплект_леопард", "44-46"],
            ["09.04", None, 2038228335633, 15, None, "Пижама_комплект_леопард", "42-44"],
            ["09.04", 11, 2039836192632, 60, None, "Пижама_брюк_леопард", "52-54"],
        ],
        title="короба",
    )


def test_boxes_sheet_format_is_recognized() -> None:
    result = parse_receive_file(_boxes_sheet_like(), supply_ref="ИНК")
    stats = result["stats"]
    assert stats["boxes_total"] == 2
    assert stats["total_qty"] == 140
    assert stats["barcodes_unique"] == 3
    assert stats["rows_dropped"] == 0
    by_code = {b["box_code"]: b for b in result["boxes"]}
    # номер короба числовой → без хвоста `.0`, иначе повторная загрузка дала бы дубли
    assert set(by_code) == {"7", "11"}
    # пустой «Номер короба» = строка-продолжение того же короба
    assert len(by_code["7"]["items"]) == 2
    assert by_code["7"]["is_mono"] is False
    assert by_code["11"]["is_mono"] is True
    # «Наименование» подхватывается — им дозаполняется справочник ШК
    assert by_code["11"]["items"][0]["name"] == "Пижама_брюк_леопард"
    # имя листа «короба» — не бренд
    assert by_code["7"]["brand"] is None


def test_numeric_box_code_loses_float_tail() -> None:
    from app.services.warehouse.packing_list import normalize_box_code

    assert normalize_box_code(7.0) == "7"
    assert normalize_box_code("7.0") == "7"
    assert normalize_box_code(7) == "7"
    assert normalize_box_code("ALT-002-ORD001-002") == "ALT-002-ORD001-002"
    assert normalize_box_code(None) == ""


def test_copy_sheets_are_skipped_to_avoid_overwriting_boxes() -> None:
    """Лист-копия опасен: номера коробов в нём те же.

    Ключ короба — его ШК, поэтому короб «7» из копии перезаписал бы короб «7» из
    основного листа. Проверено на реальном файле: там есть «короба» и
    «короба (копия)» с пересекающимися номерами.
    """
    wb = Workbook()
    main = wb.active
    main.title = "короба"
    main.append(["Номер короба", "Баркод", "Количество"])
    main.append([7, "2038228335909", 65])
    copy = wb.create_sheet("короба (копия)")
    copy.append(["Номер короба", "Баркод", "Количество"])
    copy.append([7, "2046585023346", 110])
    buf = io.BytesIO()
    wb.save(buf)

    result = parse_receive_file(buf.getvalue())
    assert result["stats"]["sheets"] == ["короба"]
    assert result["stats"]["boxes_total"] == 1
    assert result["stats"]["total_qty"] == 65


def test_duplicate_box_codes_are_reported() -> None:
    """Один номер на два разных короба — молча объединять нельзя, предупреждаем."""
    content = _xlsx(
        [
            ["Номер короба", "Баркод", "Количество"],
            [7, "1111111111111", 10],
            [8, "2222222222222", 10],
            [7, "3333333333333", 10],  # снова 7 — коллизия
        ],
        title="короба",
    )
    result = parse_receive_file(content)
    assert any("повторяющихся номеров короба" in w for w in result["warnings"])


def test_unlabelled_box_column_is_guessed_with_warning() -> None:
    """Колонка короба без подписи (в реальном файле там дата) — берём слева от баркода."""
    content = _xlsx(
        [
            [None, None, "Баркод", "Кол-во", "Наименование", "Размер"],
            [None, 1, "2046585023346", 110, "сорочка_черн", "60-62"],
            [None, 2, "2038168119416", 60, "сорочка_черн", "48-50"],
            [None, None, "2046585023346", 30, "сорочка_черн", "60-62"],
        ],
        title="короба2",
    )
    result = parse_receive_file(content)
    assert result["stats"]["boxes_total"] == 2
    assert result["stats"]["total_qty"] == 200
    assert any("не подписана" in w for w in result["warnings"])


@pytest.mark.skipif(
    not REAL_STORAGE_FILE.exists(),
    reason="реальный файл «Хранение собственный склад» недоступен",
)
def test_real_storage_file_boxes_sheet() -> None:
    """Эталон на живом файле: 52 короба, 24 баркода, 4193 шт, копия пропущена."""
    result = parse_receive_file(REAL_STORAGE_FILE.read_bytes(), supply_ref="ИНК")
    stats = result["stats"]
    assert stats["sheets"] == ["короба"], "лист-копия не должен попадать в разбор"
    assert stats["boxes_total"] == 52
    assert stats["boxes_mono"] == 29
    assert stats["barcodes_unique"] == 24
    assert stats["total_qty"] == 4193
    assert stats["rows_dropped"] == 0
