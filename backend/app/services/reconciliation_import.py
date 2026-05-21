"""Парсер XLSX-выгрузки бухгалтера для 4-way Reconciliation.

Бухгалтер выгружает из 1С / собственного учёта недельную (или месячную)
сводку. Мы парсим первый лист и ищем 6 колонок по русским названиям —
гибкость: если бухгалтер переставит колонки или переименует, мы всё равно
найдём.

Ожидаемый формат (минимум — period_from, period_to + хотя бы revenue):

    | Период с   | Период по  | Выручка ₽ | Возвраты | Комиссия | К выплате |
    | 2026-05-05 | 2026-05-11 | 1 234 567 | 50 000   | 250 000  | 950 000   |
    | 2026-05-12 | 2026-05-18 |   985 000 | 30 000   | 200 000  | 750 000   |

Заголовки могут быть в любом регистре, любым языком (RU/EN), с/без знаков
препинания. Парсер нормализует строку (lower, без пробелов/знаков) и
матчит по подстрокам.

Возвращает массив `ParsedRow(period_from, period_to, revenue_gross_rub,
revenue_returns_rub, commission_rub, payout_rub)` + список ошибок.

MVP-парсер: один лист, не обрабатывает merged-cells.
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook


@dataclass
class ParsedRow:
    period_from: date
    period_to: date
    revenue_gross_rub: Decimal | None = None
    revenue_returns_rub: Decimal | None = None
    commission_rub: Decimal | None = None
    payout_rub: Decimal | None = None


@dataclass
class ParseResult:
    rows: list[ParsedRow]
    errors: list[str]
    sheet_name: str
    header_row: int


# Кандидаты заголовков — нормализованные (lower, alpha+digit only).
# Найдём по substring-match: для "period_from" подходят "период с", "periodfrom",
# "perfrom", "от", "with" — главное чтобы было одно из ключевых слов в колонке.
_FIELD_KEYWORDS: dict[str, list[str]] = {
    "period_from": ["периодс", "сдат", "датасн", "starts", "fromdate", "periodfrom", "weekstart"],
    "period_to": ["периодпо", "подат", "datedend", "ends", "todate", "periodto", "weekend"],
    "revenue_gross_rub": ["выручка", "продажи", "revenue", "gross", "оборот"],
    "revenue_returns_rub": ["возврат", "returns"],
    "commission_rub": ["комиссия", "commission"],
    "payout_rub": ["квыплате", "payout", "перечислению", "квыпл"],
}


def _normalize(s: str) -> str:
    """Приводим заголовок к lower + только буквы/цифры — для loose-match."""
    return re.sub(r"[^a-zа-я0-9]+", "", s.lower())


def _match_field(cell_value: Any) -> str | None:
    """Найти к какому полю относится заголовок, или None."""
    if cell_value is None:
        return None
    normalized = _normalize(str(cell_value))
    if not normalized:
        return None
    for field, keywords in _FIELD_KEYWORDS.items():
        for kw in keywords:
            if kw in normalized:
                return field
    return None


def _to_date(v: Any) -> date | None:
    """Преобразовать значение ячейки в date. Поддерживает datetime, date, ISO-строку."""
    if v is None:
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y.%m.%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
    return None


def _to_decimal(v: Any) -> Decimal | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float, Decimal)):
        try:
            return Decimal(str(v))
        except InvalidOperation:
            return None
    if isinstance(v, str):
        # «1 234 567,89» / «1,234,567.89» / «1 234.56» → нормализация
        cleaned = v.replace("\xa0", " ").replace(" ", "")
        # Если есть и запятая и точка — точка thousand sep, запятая decimal:
        if "," in cleaned and "." in cleaned:
            cleaned = cleaned.replace(".", "").replace(",", ".")
        elif "," in cleaned:
            cleaned = cleaned.replace(",", ".")
        try:
            return Decimal(cleaned)
        except InvalidOperation:
            return None
    return None


def parse_bookkeeper_xlsx(file_bytes: bytes) -> ParseResult:
    """Распарсить XLSX от бухгалтера. Возвращает rows + список ошибок.

    Алгоритм:
        1. Открыть первый лист (или единственный, если имя 'Sheet1')
        2. Найти строку с заголовками: первая строка где >= 2 ячеек матчат FIELD_KEYWORDS
        3. Запомнить mapping `{column_index: field_name}`
        4. Итерация по строкам ниже: собрать ParsedRow, скип если нет period_from/to
        5. Вернуть rows + errors (про скип-строки)
    """
    errors: list[str] = []
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        return ParseResult(rows=[], errors=[f"Не удалось открыть XLSX: {e}"], sheet_name="", header_row=0)

    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    # Поиск header row — пробуем первые 10 строк
    header_map: dict[int, str] = {}
    header_row = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        candidates: dict[int, str] = {}
        for col_idx, val in enumerate(row):
            field = _match_field(val)
            if field:
                candidates[col_idx] = field
        if len(candidates) >= 2:  # минимум 2 матча = это header
            header_map = candidates
            header_row = row_idx
            break

    if not header_map:
        return ParseResult(
            rows=[],
            errors=[
                "Не найдены колонки заголовка. Ожидаются (любым языком): "
                "Период с / Период по / Выручка / Возвраты / Комиссия / К выплате"
            ],
            sheet_name=sheet_name,
            header_row=0,
        )

    rows: list[ParsedRow] = []
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        # Собрать значения по полям
        data: dict[str, Any] = {}
        for col_idx, field in header_map.items():
            if col_idx < len(row):
                data[field] = row[col_idx]

        pf = _to_date(data.get("period_from"))
        pt = _to_date(data.get("period_to"))
        if not pf or not pt:
            # Пустая строка или некорректный период — скип
            if any(data.get(k) for k in ("revenue_gross_rub", "commission_rub")):
                errors.append(
                    f"Строка {row_idx}: пропущена — некорректные даты периода "
                    f"({data.get('period_from')!r} … {data.get('period_to')!r})"
                )
            continue

        rows.append(ParsedRow(
            period_from=pf,
            period_to=pt,
            revenue_gross_rub=_to_decimal(data.get("revenue_gross_rub")),
            revenue_returns_rub=_to_decimal(data.get("revenue_returns_rub")),
            commission_rub=_to_decimal(data.get("commission_rub")),
            payout_rub=_to_decimal(data.get("payout_rub")),
        ))

    return ParseResult(rows=rows, errors=errors, sheet_name=sheet_name, header_row=header_row)
