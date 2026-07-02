"""Парсинг банковских выписок (TASK-DEV-093).

Форматы:
- 1С «1CClientBankExchange» (.txt) — стандарт обмена банк-клиентов всех
  российских банков (ТБанк/ВТБ/Сбер…). Кодировка почти всегда windows-1251 —
  декод-каскад utf-8-sig → utf-8 → cp1251, валидация по маркеру.
- Excel/CSV — наш шаблон (или произвольные колонки + пользовательский маппинг).

Выход — список нормализованных строк-операций:
  {op_date, op_kind (income|expense), amount, counterparty, raw_description,
   doc_number, account_number?}
Дедуп — sha256(account_id|date|amount|doc_number|normalized(description))[:32].
"""
from __future__ import annotations

import csv
import hashlib
import io
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

MARKER_1C = "1CClientBankExchange"

# Колонки нашего Excel/CSV-шаблона → поля операции.
TEMPLATE_COLUMNS = {
    "Дата": "op_date",
    "Тип": "op_kind",  # Доход | Расход
    "Сумма": "amount",
    "Счёт": "account_name",
    "Статья": "article_name",
    "Контрагент": "counterparty",
    "Назначение платежа": "raw_description",
    "№ документа": "doc_number",
}

_KIND_ALIASES = {
    "доход": "income", "income": "income", "поступление": "income", "приход": "income",
    "расход": "expense", "expense": "expense", "списание": "expense",
}


class StatementParseError(ValueError):
    pass


def decode_statement(raw: bytes) -> str:
    """utf-8-sig → utf-8 → cp1251. 1С-файлы почти всегда windows-1251."""
    for enc in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("cp1251", errors="replace")


def detect_format(filename: str, raw: bytes) -> str:
    """1c | excel | csv по содержимому/расширению."""
    head = raw[:4096]
    if raw[:2] == b"PK":  # xlsx = zip
        return "excel"
    try:
        text_head = decode_statement(head)
    except Exception:
        text_head = ""
    if MARKER_1C in text_head:
        return "1c"
    lower = (filename or "").lower()
    if lower.endswith((".xlsx", ".xls")):
        return "excel"
    if lower.endswith(".csv"):
        return "csv"
    if lower.endswith(".txt"):
        # txt без маркера — считаем 1С-попыткой, парсер отдаст понятную ошибку
        return "1c"
    return "csv"


def _parse_1c_date(v: str) -> date | None:
    v = (v or "").strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(v, fmt).date()
        except ValueError:
            continue
    return None


def _to_amount(v: Any) -> Decimal | None:
    if v is None:
        return None
    s = str(v).strip().replace(" ", "").replace("\xa0", "").replace(",", ".")
    if not s:
        return None
    try:
        return Decimal(s).quantize(Decimal("0.01"))
    except InvalidOperation:
        return None


def parse_1c(raw: bytes) -> dict[str, Any]:
    """Парсер 1CClientBankExchange.

    Возвращает {"our_accounts": [номера р/с из СекцияРасчСчет],
                "rows": [операции]}. Направление операции: если
    ПлательщикРасчСчет — наш счёт → expense; ПолучательРасчСчет — наш → income.
    """
    text = decode_statement(raw)
    if MARKER_1C not in text[:1000]:
        raise StatementParseError(
            "Файл не похож на выписку 1С (нет маркера 1CClientBankExchange)"
        )

    our_accounts: set[str] = set()
    rows: list[dict[str, Any]] = []
    doc: dict[str, str] | None = None
    in_account_section = False

    for line_raw in text.splitlines():
        line = line_raw.strip("﻿ \t\r\n")
        if not line:
            continue
        upper = line.split("=", 1)[0]
        if line.startswith("СекцияРасчСчет"):
            in_account_section = True
            continue
        if line.startswith("КонецРасчСчет"):
            in_account_section = False
            continue
        if line.startswith("СекцияДокумент"):
            doc = {}
            continue
        if line.startswith("КонецДокумента"):
            if doc is not None:
                row = _doc_to_row(doc, our_accounts)
                if row is not None:
                    rows.append(row)
            doc = None
            continue
        if "=" not in line:
            continue
        key, value = line.split("=", 1)
        key, value = key.strip(), value.strip()
        if in_account_section and key == "РасчСчет":
            our_accounts.add(value)
        elif doc is not None:
            doc[key] = value

    if not rows and not our_accounts:
        raise StatementParseError("В выписке не найдено ни счетов, ни документов")
    return {"our_accounts": sorted(our_accounts), "rows": rows}


def _doc_to_row(doc: dict[str, str], our_accounts: set[str]) -> dict[str, Any] | None:
    amount = _to_amount(doc.get("Сумма"))
    if amount is None or amount == 0:
        return None
    payer_acc = doc.get("ПлательщикРасчСчет", "").strip()
    payee_acc = doc.get("ПолучательРасчСчет", "").strip()

    # ДатаСписано (расход) / ДатаПоступило (доход) точнее, чем Дата документа.
    op_date = (
        _parse_1c_date(doc.get("ДатаСписано", ""))
        or _parse_1c_date(doc.get("ДатаПоступило", ""))
        or _parse_1c_date(doc.get("Дата", ""))
    )
    if op_date is None:
        return None

    if payer_acc and payer_acc in our_accounts:
        op_kind = "expense"
        counterparty = doc.get("Получатель1") or doc.get("Получатель") or ""
        account_number = payer_acc
    elif payee_acc and payee_acc in our_accounts:
        op_kind = "income"
        counterparty = doc.get("Плательщик1") or doc.get("Плательщик") or ""
        account_number = payee_acc
    else:
        # Ни один счёт не наш (или секции РасчСчет нет) — определяем по
        # наличию ДатаСписано/ДатаПоступило.
        if doc.get("ДатаСписано"):
            op_kind = "expense"
            counterparty = doc.get("Получатель1") or doc.get("Получатель") or ""
            account_number = payer_acc
        elif doc.get("ДатаПоступило"):
            op_kind = "income"
            counterparty = doc.get("Плательщик1") or doc.get("Плательщик") or ""
            account_number = payee_acc
        else:
            return None

    return {
        "op_date": op_date.isoformat(),
        "op_kind": op_kind,
        "amount": float(amount),
        "counterparty": counterparty.strip()[:255] or None,
        "raw_description": (doc.get("НазначениеПлатежа") or "").strip() or None,
        "doc_number": (doc.get("Номер") or "").strip()[:64] or None,
        "account_number": account_number or None,
    }


def parse_tabular(
    raw: bytes, *, file_format: str, mapping: dict[str, str] | None = None
) -> dict[str, Any]:
    """Excel/CSV: без mapping — вернуть колонки и предположение маппинга
    (needs_mapping, если не все обязательные поля распознаны); с mapping —
    распарсить строки. mapping: {"<заголовок колонки>": "<поле операции>"}."""
    if file_format == "excel":
        headers, records = _read_excel(raw)
    else:
        headers, records = _read_csv(raw)

    suggest: dict[str, str] = {}
    for h in headers:
        hn = (h or "").strip()
        if hn in TEMPLATE_COLUMNS:
            suggest[hn] = TEMPLATE_COLUMNS[hn]
        else:
            low = hn.lower()
            if "дат" in low:
                suggest.setdefault(hn, "op_date")
            elif "сумм" in low:
                suggest.setdefault(hn, "amount")
            elif "тип" in low or "вид" in low:
                suggest.setdefault(hn, "op_kind")
            elif "контрагент" in low or "плательщик" in low or "получатель" in low:
                suggest.setdefault(hn, "counterparty")
            elif "назнач" in low or "коммент" in low or "описан" in low:
                suggest.setdefault(hn, "raw_description")
            elif "стать" in low or "категор" in low:
                suggest.setdefault(hn, "article_name")
            elif "счет" in low or "счёт" in low:
                suggest.setdefault(hn, "account_name")
            elif "номер" in low or "док" in low:
                suggest.setdefault(hn, "doc_number")

    eff_mapping = mapping or suggest
    required = {"op_date", "amount"}
    if not required.issubset(set(eff_mapping.values())):
        return {
            "needs_mapping": True,
            "columns": headers,
            "mapping_suggest": suggest,
            "rows": [],
        }

    col_to_field = {h: f for h, f in eff_mapping.items() if f}
    rows: list[dict[str, Any]] = []
    for rec in records:
        item: dict[str, Any] = {}
        for h, f in col_to_field.items():
            item[f] = rec.get(h)
        amount = _to_amount(item.get("amount"))
        d = item.get("op_date")
        if isinstance(d, datetime):
            op_date = d.date()
        elif isinstance(d, date):
            op_date = d
        else:
            op_date = _parse_1c_date(str(d or ""))
        if amount is None or op_date is None:
            continue
        kind_raw = str(item.get("op_kind") or "").strip().lower()
        op_kind = _KIND_ALIASES.get(kind_raw)
        if op_kind is None:
            # знак суммы: отрицательная → расход
            op_kind = "expense" if amount < 0 else "income"
        rows.append(
            {
                "op_date": op_date.isoformat(),
                "op_kind": op_kind,
                "amount": float(abs(amount)),
                "counterparty": (str(item.get("counterparty") or "").strip()[:255] or None),
                "raw_description": (str(item.get("raw_description") or "").strip() or None),
                "doc_number": (str(item.get("doc_number") or "").strip()[:64] or None),
                "article_name": (str(item.get("article_name") or "").strip() or None),
                "account_name": (str(item.get("account_name") or "").strip() or None),
            }
        )
    return {
        "needs_mapping": False,
        "columns": headers,
        "mapping_suggest": suggest,
        "rows": rows,
    }


def _read_excel(raw: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        raise StatementParseError("Пустой Excel-файл")
    records = []
    for row in rows_iter:
        if row is None or all(v is None for v in row):
            continue
        records.append({headers[i]: row[i] for i in range(min(len(headers), len(row)))})
    wb.close()
    return headers, records


def _read_csv(raw: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    text = decode_statement(raw)
    sample = text[:2048]
    delimiter = ";" if sample.count(";") >= sample.count(",") else ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    headers = [h.strip() for h in (reader.fieldnames or [])]
    records = [{(k or "").strip(): v for k, v in rec.items()} for rec in reader]
    return headers, records


_WS_RE = re.compile(r"\s+")


def dedup_hash(
    *,
    account_id: int | None,
    op_date: str,
    amount: float,
    doc_number: str | None,
    raw_description: str | None,
) -> str:
    normalized = _WS_RE.sub(" ", (raw_description or "").strip().lower())
    base = f"{account_id or 0}|{op_date}|{amount:.2f}|{doc_number or ''}|{normalized}"
    return hashlib.sha256(base.encode("utf-8")).hexdigest()[:32]
