"""Excel import/export for reference-data tables (справочники).

Each entity declares its column schema (a list of header names that are also
attribute names on the ORM model). The same list is used for both export and
import — a round-trip works as long as headers are preserved (column order can
change; extra columns are ignored).

Per-entity import strategy:
- products             — upsert by nm_id (PK).
- cogs                 — upsert by (nm_id, valid_from); creates parent product if missing.
- opex_categories      — upsert by name (unique).
- opex_entries         — upsert by id; resolves category_name → category_id.
- artificial_orders    — upsert by id.
- external_ad_costs    — upsert by id.
- sales_plans          — upsert by (year, month, scope_type, scope_id) (uq index).
- wb_tariff_categories — upsert by name (unique).
- settings             — upsert by key (PK).

Bad rows are reported in `errors` (with row number) and skipped; the rest is
committed atomically.
"""
from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Awaitable, Callable

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.services.tenant_context import get_tenant

from app.db.models import (
    AppSetting,
    ArtificialOrder,
    Cogs,
    ExternalAdCost,
    OffPlatformStockMovement,
    OpexCategory,
    OpexEntry,
    Product,
    ProductGroup,
    ProductGroupAssignment,
    SalesPlan,
    SettingTimeline,
    Supply,
    WbTariffCategory,
)


# ---------------------------------------------------------------------------
# Per-entity column schemas (header == ORM attr name when no remap is needed)
# ---------------------------------------------------------------------------

SCHEMAS: dict[str, list[str]] = {
    "products": ["nm_id", "vendor_code", "subject", "brand", "category", "is_archived"],
    "cogs": ["nm_id", "valid_from", "cost_rub", "packaging_rub", "fulfillment_rub"],
    "opex_categories": ["name", "kind", "is_fixed", "in_operating", "cf_section"],
    "opex_entries": ["id", "entry_date", "category_name", "amount", "contractor", "comment"],
    "artificial_orders": [
        "id", "type", "order_dt", "completion_dt",
        "nm_id", "qty", "gross_amount", "contractor_fee", "comment",
    ],
    "external_ad_costs": ["id", "spend_date", "end_date", "nm_id", "channel", "amount", "comment"],
    "sales_plans": [
        "period_year", "period_month", "scope_type", "scope_id",
        "planned_orders_qty", "planned_orders_revenue",
        "planned_sales_qty", "planned_sales_revenue",
        "planned_profit", "planned_marketing_cost",
        "comment",
    ],
    "wb_tariff_categories": [
        "name", "commission_pct", "default_logistics_per_unit", "sort_order",
    ],
    "settings": ["key", "value"],
    "setting_timeline": ["id", "key", "effective_from", "value", "comment"],
    "off_platform_stock": [
        "id", "dt", "nm_id", "kind", "qty", "unit_cost", "comment",
    ],
    "product_groups": ["id", "name", "manager_name", "color", "comment"],
    "product_group_assignments": ["group_name", "nm_id"],
    "supplies": [
        "id", "nm_id", "vendor_code", "supply_date",
        "qty", "cost_per_unit", "currency",
        "vendor", "invoice_number",
        "paid_status", "paid_date", "paid_amount",
        "notes",
    ],
    "jam_queries": [
        "nm_id", "query", "period_start", "period_end",
        "orders", "clicks", "views", "ad_spent",
    ],
}

ENTITY_LABELS: dict[str, str] = {
    "products": "Товары",
    "cogs": "Себестоимость (история)",
    "opex_categories": "Категории OPEX",
    "opex_entries": "Записи OPEX",
    "artificial_orders": "Корректировки выручки",
    "external_ad_costs": "Внеш. маркетинг",
    "sales_plans": "Планы (План-Факт)",
    "wb_tariff_categories": "Категории WB (тарифы)",
    "settings": "Настройки (KV)",
    "setting_timeline": "Расписание налогов / НДС",
    "off_platform_stock": "Off-WB склад (движения)",
    "product_groups": "Группы товаров",
    "product_group_assignments": "Привязка SKU к группам",
    "supplies": "Закупки у поставщиков",
    "jam_queries": "Джем — поисковые запросы",
}


# ---------------------------------------------------------------------------
# Coercions: openpyxl returns native Python types (str/int/float/datetime/None).
# ---------------------------------------------------------------------------


def _to_str(v: Any) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _to_int(v: Any) -> int | None:
    if v in (None, ""):
        return None
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, (int, float, Decimal)):
        return int(v)
    s = str(v).strip().replace(" ", "")
    if not s:
        return None
    return int(float(s))


def _to_decimal(v: Any) -> Decimal | None:
    if v in (None, ""):
        return None
    if isinstance(v, Decimal):
        return v
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    s = str(v).strip().replace(" ", "").replace(",", ".")
    if not s:
        return None
    try:
        return Decimal(s)
    except InvalidOperation as e:
        raise ValueError(f"не число: {v!r}") from e


def _to_date(v: Any) -> date | None:
    if v in (None, ""):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    try:
        return date.fromisoformat(s[:10])
    except ValueError as e:
        raise ValueError(f"не дата (нужно YYYY-MM-DD): {v!r}") from e


def _to_bool(v: Any) -> bool:
    if v in (None, ""):
        return False
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v != 0
    return str(v).strip().lower() in ("1", "true", "yes", "да", "y", "t", "истина")


# ---------------------------------------------------------------------------
# Per-entity export & import
# ---------------------------------------------------------------------------


async def _export_products(session: AsyncSession) -> list[list[Any]]:
    rows = (await session.execute(select(Product).order_by(Product.nm_id))).scalars().all()
    return [
        [p.nm_id, p.vendor_code, p.subject, p.brand, p.category, p.is_archived]
        for p in rows
    ]


async def _import_products(
    session: AsyncSession, rows: list[dict[str, Any]]
) -> dict[str, Any]:
    inserted = updated = skipped = 0
    errors: list[str] = []
    for i, r in enumerate(rows, start=2):
        try:
            nm_id = _to_int(r.get("nm_id"))
            if not nm_id:
                skipped += 1
                continue
            existing = await session.get(Product, nm_id)
            if existing:
                if (vc := _to_str(r.get("vendor_code"))) is not None:
                    existing.vendor_code = vc
                if (s := _to_str(r.get("subject"))) is not None:
                    existing.subject = s
                if (b := _to_str(r.get("brand"))) is not None:
                    existing.brand = b
                if (c := _to_str(r.get("category"))) is not None:
                    existing.category = c
                if "is_archived" in r and r["is_archived"] is not None:
                    existing.is_archived = _to_bool(r["is_archived"])
                updated += 1
            else:
                session.add(Product(
                    nm_id=nm_id,
                    vendor_code=_to_str(r.get("vendor_code")),
                    subject=_to_str(r.get("subject")),
                    brand=_to_str(r.get("brand")),
                    category=_to_str(r.get("category")),
                    is_archived=_to_bool(r.get("is_archived")),
                ))
                inserted += 1
        except Exception as e:
            errors.append(f"строка {i}: {e}")
            skipped += 1
    return {"inserted": inserted, "updated": updated, "skipped": skipped, "errors": errors}


async def _export_cogs(session: AsyncSession) -> list[list[Any]]:
    rows = (
        await session.execute(select(Cogs).order_by(Cogs.nm_id, Cogs.valid_from))
    ).scalars().all()
    return [[c.nm_id, c.valid_from, c.cost_rub, c.packaging_rub, c.fulfillment_rub] for c in rows]


async def _import_cogs(
    session: AsyncSession, rows: list[dict[str, Any]]
) -> dict[str, Any]:
    inserted = updated = skipped = 0
    errors: list[str] = []
    for i, r in enumerate(rows, start=2):
        try:
            nm_id = _to_int(r.get("nm_id"))
            if not nm_id:
                skipped += 1
                continue
            valid_from = _to_date(r.get("valid_from")) or date.today()
            cost = _to_decimal(r.get("cost_rub")) or Decimal(0)
            pack = _to_decimal(r.get("packaging_rub")) or Decimal(0)
            ful = _to_decimal(r.get("fulfillment_rub")) or Decimal(0)
            await session.execute(
                pg_insert(Product)
                .values(tenant_id=get_tenant(session), nm_id=nm_id)
                .on_conflict_do_nothing(index_elements=["nm_id"])
            )
            existing = (await session.execute(
                select(Cogs).where(Cogs.nm_id == nm_id, Cogs.valid_from == valid_from)
            )).scalars().first()
            if existing:
                existing.cost_rub = cost
                existing.packaging_rub = pack
                existing.fulfillment_rub = ful
                updated += 1
            else:
                session.add(Cogs(
                    nm_id=nm_id, valid_from=valid_from,
                    cost_rub=cost, packaging_rub=pack, fulfillment_rub=ful,
                ))
                inserted += 1
        except Exception as e:
            errors.append(f"строка {i}: {e}")
            skipped += 1
    return {"inserted": inserted, "updated": updated, "skipped": skipped, "errors": errors}


async def _export_opex_categories(session: AsyncSession) -> list[list[Any]]:
    rows = (await session.execute(
        select(OpexCategory).order_by(OpexCategory.kind, OpexCategory.name)
    )).scalars().all()
    return [[c.name, c.kind, c.is_fixed, c.in_operating, c.cf_section] for c in rows]


async def _import_opex_categories(
    session: AsyncSession, rows: list[dict[str, Any]]
) -> dict[str, Any]:
    inserted = updated = skipped = 0
    errors: list[str] = []
    for i, r in enumerate(rows, start=2):
        try:
            name = _to_str(r.get("name"))
            if not name:
                skipped += 1
                continue
            kind = (_to_str(r.get("kind")) or "expense").lower()
            if kind not in ("expense", "income"):
                kind = "expense"
            cf_section = (_to_str(r.get("cf_section")) or "operating").lower()
            if cf_section not in ("operating", "investing", "financing"):
                cf_section = "operating"
            existing = (await session.execute(
                select(OpexCategory).where(OpexCategory.name == name)
            )).scalars().first()
            kwargs = dict(
                name=name, kind=kind,
                is_fixed=_to_bool(r.get("is_fixed")),
                in_operating=_to_bool(r.get("in_operating")),
                cf_section=cf_section,
            )
            if existing:
                for k, v in kwargs.items():
                    setattr(existing, k, v)
                updated += 1
            else:
                session.add(OpexCategory(**kwargs))
                inserted += 1
        except Exception as e:
            errors.append(f"строка {i}: {e}")
            skipped += 1
    return {"inserted": inserted, "updated": updated, "skipped": skipped, "errors": errors}


async def _export_opex_entries(session: AsyncSession) -> list[list[Any]]:
    rows = (await session.execute(
        select(OpexEntry, OpexCategory.name)
        .join(OpexCategory, OpexEntry.category_id == OpexCategory.id)
        .order_by(OpexEntry.entry_date, OpexEntry.id)
    )).all()
    return [[e.id, e.entry_date, name, e.amount, e.contractor, e.comment] for e, name in rows]


async def _import_opex_entries(
    session: AsyncSession, rows: list[dict[str, Any]]
) -> dict[str, Any]:
    inserted = updated = skipped = 0
    errors: list[str] = []
    cats = (await session.execute(select(OpexCategory.id, OpexCategory.name))).all()
    cat_by_name: dict[str, int] = {row.name: row.id for row in cats}
    for i, r in enumerate(rows, start=2):
        try:
            cat_name = _to_str(r.get("category_name"))
            if not cat_name or cat_name not in cat_by_name:
                errors.append(f"строка {i}: неизвестная категория {cat_name!r}")
                skipped += 1
                continue
            entry_date = _to_date(r.get("entry_date"))
            amount = _to_decimal(r.get("amount"))
            if not entry_date or amount is None:
                errors.append(f"строка {i}: пустая дата или сумма")
                skipped += 1
                continue
            row_id = _to_int(r.get("id"))
            existing = await session.get(OpexEntry, row_id) if row_id else None
            kwargs = dict(
                entry_date=entry_date,
                category_id=cat_by_name[cat_name],
                amount=amount,
                contractor=_to_str(r.get("contractor")),
                comment=_to_str(r.get("comment")),
            )
            if existing:
                for k, v in kwargs.items():
                    setattr(existing, k, v)
                updated += 1
            else:
                session.add(OpexEntry(**kwargs))
                inserted += 1
        except Exception as e:
            errors.append(f"строка {i}: {e}")
            skipped += 1
    return {"inserted": inserted, "updated": updated, "skipped": skipped, "errors": errors}


async def _export_artif(session: AsyncSession) -> list[list[Any]]:
    rows = (await session.execute(
        select(ArtificialOrder).order_by(ArtificialOrder.order_dt, ArtificialOrder.id)
    )).scalars().all()
    return [[
        a.id, a.type, a.order_dt, a.completion_dt, a.nm_id, a.qty,
        a.gross_amount, a.contractor_fee, a.comment,
    ] for a in rows]


async def _import_artif(
    session: AsyncSession, rows: list[dict[str, Any]]
) -> dict[str, Any]:
    inserted = updated = skipped = 0
    errors: list[str] = []
    for i, r in enumerate(rows, start=2):
        try:
            type_ = _to_str(r.get("type")) or ""
            order_dt = _to_date(r.get("order_dt"))
            if not type_ or not order_dt:
                errors.append(f"строка {i}: нужны type и order_dt")
                skipped += 1
                continue
            row_id = _to_int(r.get("id"))
            existing = await session.get(ArtificialOrder, row_id) if row_id else None
            kwargs = dict(
                type=type_,
                order_dt=order_dt,
                completion_dt=_to_date(r.get("completion_dt")),
                nm_id=_to_int(r.get("nm_id")),
                qty=_to_int(r.get("qty")) or 1,
                gross_amount=_to_decimal(r.get("gross_amount")) or Decimal(0),
                contractor_fee=_to_decimal(r.get("contractor_fee")) or Decimal(0),
                comment=_to_str(r.get("comment")),
            )
            if existing:
                for k, v in kwargs.items():
                    setattr(existing, k, v)
                updated += 1
            else:
                session.add(ArtificialOrder(**kwargs))
                inserted += 1
        except Exception as e:
            errors.append(f"строка {i}: {e}")
            skipped += 1
    return {"inserted": inserted, "updated": updated, "skipped": skipped, "errors": errors}


async def _export_ext_ad(session: AsyncSession) -> list[list[Any]]:
    rows = (await session.execute(
        select(ExternalAdCost).order_by(ExternalAdCost.spend_date, ExternalAdCost.id)
    )).scalars().all()
    return [[r.id, r.spend_date, r.end_date, r.nm_id, r.channel, r.amount, r.comment] for r in rows]


async def _import_ext_ad(
    session: AsyncSession, rows: list[dict[str, Any]]
) -> dict[str, Any]:
    inserted = updated = skipped = 0
    errors: list[str] = []
    for i, r in enumerate(rows, start=2):
        try:
            spend_date = _to_date(r.get("spend_date"))
            channel = _to_str(r.get("channel")) or "other"
            amount = _to_decimal(r.get("amount"))
            if not spend_date or amount is None:
                errors.append(f"строка {i}: нужны spend_date и amount")
                skipped += 1
                continue
            row_id = _to_int(r.get("id"))
            existing = await session.get(ExternalAdCost, row_id) if row_id else None
            kwargs = dict(
                spend_date=spend_date,
                end_date=_to_date(r.get("end_date")),
                nm_id=_to_int(r.get("nm_id")),
                channel=channel,
                amount=amount,
                comment=_to_str(r.get("comment")),
            )
            if existing:
                for k, v in kwargs.items():
                    setattr(existing, k, v)
                updated += 1
            else:
                session.add(ExternalAdCost(**kwargs))
                inserted += 1
        except Exception as e:
            errors.append(f"строка {i}: {e}")
            skipped += 1
    return {"inserted": inserted, "updated": updated, "skipped": skipped, "errors": errors}


async def _export_plans(session: AsyncSession) -> list[list[Any]]:
    rows = (await session.execute(
        select(SalesPlan).order_by(SalesPlan.period_year, SalesPlan.period_month, SalesPlan.id)
    )).scalars().all()
    return [[
        p.period_year, p.period_month, p.scope_type, p.scope_id,
        p.planned_orders_qty, p.planned_orders_revenue,
        p.planned_sales_qty, p.planned_sales_revenue,
        p.planned_profit, p.planned_marketing_cost, p.comment,
    ] for p in rows]


async def _import_plans(
    session: AsyncSession, rows: list[dict[str, Any]]
) -> dict[str, Any]:
    inserted = updated = skipped = 0
    errors: list[str] = []
    for i, r in enumerate(rows, start=2):
        try:
            year = _to_int(r.get("period_year"))
            month = _to_int(r.get("period_month"))
            if not year or not month:
                errors.append(f"строка {i}: нужны period_year и period_month")
                skipped += 1
                continue
            scope_type = _to_str(r.get("scope_type")) or "store"
            scope_id = _to_int(r.get("scope_id"))
            existing = (await session.execute(
                select(SalesPlan).where(
                    SalesPlan.period_year == year,
                    SalesPlan.period_month == month,
                    SalesPlan.scope_type == scope_type,
                    SalesPlan.scope_id == scope_id,
                )
            )).scalars().first()
            kwargs = dict(
                period_year=year, period_month=month,
                scope_type=scope_type, scope_id=scope_id,
                planned_orders_qty=_to_int(r.get("planned_orders_qty")) or 0,
                planned_orders_revenue=_to_decimal(r.get("planned_orders_revenue")) or Decimal(0),
                planned_sales_qty=_to_int(r.get("planned_sales_qty")) or 0,
                planned_sales_revenue=_to_decimal(r.get("planned_sales_revenue")) or Decimal(0),
                planned_profit=_to_decimal(r.get("planned_profit")) or Decimal(0),
                planned_marketing_cost=_to_decimal(r.get("planned_marketing_cost")) or Decimal(0),
                comment=_to_str(r.get("comment")),
            )
            if existing:
                for k, v in kwargs.items():
                    setattr(existing, k, v)
                updated += 1
            else:
                session.add(SalesPlan(**kwargs))
                inserted += 1
        except Exception as e:
            errors.append(f"строка {i}: {e}")
            skipped += 1
    return {"inserted": inserted, "updated": updated, "skipped": skipped, "errors": errors}


async def _export_tariffs(session: AsyncSession) -> list[list[Any]]:
    rows = (await session.execute(
        select(WbTariffCategory).order_by(
            WbTariffCategory.sort_order, WbTariffCategory.name
        )
    )).scalars().all()
    return [[
        r.name, r.commission_pct, r.default_logistics_per_unit, r.sort_order,
    ] for r in rows]


async def _import_tariffs(
    session: AsyncSession, rows: list[dict[str, Any]]
) -> dict[str, Any]:
    inserted = updated = skipped = 0
    errors: list[str] = []
    for i, r in enumerate(rows, start=2):
        try:
            name = _to_str(r.get("name"))
            if not name:
                skipped += 1
                continue
            existing = (await session.execute(
                select(WbTariffCategory).where(WbTariffCategory.name == name)
            )).scalars().first()
            kwargs = dict(
                name=name,
                commission_pct=_to_decimal(r.get("commission_pct")) or Decimal(18),
                default_logistics_per_unit=
                    _to_decimal(r.get("default_logistics_per_unit")) or Decimal(80),
                sort_order=_to_int(r.get("sort_order")) or 0,
            )
            if existing:
                for k, v in kwargs.items():
                    setattr(existing, k, v)
                updated += 1
            else:
                session.add(WbTariffCategory(**kwargs))
                inserted += 1
        except Exception as e:
            errors.append(f"строка {i}: {e}")
            skipped += 1
    return {"inserted": inserted, "updated": updated, "skipped": skipped, "errors": errors}


async def _export_settings(session: AsyncSession) -> list[list[Any]]:
    rows = (await session.execute(
        select(AppSetting).order_by(AppSetting.key)
    )).scalars().all()
    return [[r.key, r.value] for r in rows]


async def _import_settings(
    session: AsyncSession, rows: list[dict[str, Any]]
) -> dict[str, Any]:
    inserted = updated = skipped = 0
    errors: list[str] = []
    for i, r in enumerate(rows, start=2):
        try:
            key = _to_str(r.get("key"))
            if not key:
                skipped += 1
                continue
            value = r.get("value")
            value_str = "" if value is None else str(value)
            existing = await session.get(AppSetting, key)
            if existing:
                existing.value = value_str
                updated += 1
            else:
                session.add(AppSetting(key=key, value=value_str))
                inserted += 1
        except Exception as e:
            errors.append(f"строка {i}: {e}")
            skipped += 1
    return {"inserted": inserted, "updated": updated, "skipped": skipped, "errors": errors}


# ---------------------------------------------------------------------------
# Registry & top-level helpers
# ---------------------------------------------------------------------------

ExportFn = Callable[[AsyncSession], Awaitable[list[list[Any]]]]
ImportFn = Callable[[AsyncSession, list[dict[str, Any]]], Awaitable[dict[str, Any]]]

async def _export_setting_timeline(session: AsyncSession) -> list[list[Any]]:
    rows = (await session.execute(
        select(SettingTimeline).order_by(
            SettingTimeline.key, SettingTimeline.effective_from
        )
    )).scalars().all()
    return [[r.id, r.key, r.effective_from, r.value, r.comment] for r in rows]


async def _import_setting_timeline(
    session: AsyncSession, rows: list[dict[str, Any]]
) -> dict[str, Any]:
    inserted = updated = skipped = 0
    errors: list[str] = []
    for i, r in enumerate(rows, start=2):
        try:
            key = _to_str(r.get("key"))
            effective_from = _to_date(r.get("effective_from"))
            if not key or not effective_from:
                errors.append(f"строка {i}: нужны key и effective_from")
                skipped += 1
                continue
            value = _to_str(r.get("value"))
            existing = (await session.execute(
                select(SettingTimeline).where(
                    SettingTimeline.key == key,
                    SettingTimeline.effective_from == effective_from,
                )
            )).scalars().first()
            if existing:
                existing.value = value
                existing.comment = _to_str(r.get("comment"))
                updated += 1
            else:
                session.add(SettingTimeline(
                    key=key,
                    effective_from=effective_from,
                    value=value,
                    comment=_to_str(r.get("comment")),
                ))
                inserted += 1
        except Exception as e:
            errors.append(f"строка {i}: {e}")
            skipped += 1
    return {"inserted": inserted, "updated": updated, "skipped": skipped, "errors": errors}


async def _export_off_platform(session: AsyncSession) -> list[list[Any]]:
    rows = (await session.execute(
        select(OffPlatformStockMovement).order_by(
            OffPlatformStockMovement.dt, OffPlatformStockMovement.id
        )
    )).scalars().all()
    return [[
        m.id, m.dt, m.nm_id, m.kind, m.qty, m.unit_cost, m.comment,
    ] for m in rows]


async def _import_off_platform(
    session: AsyncSession, rows: list[dict[str, Any]]
) -> dict[str, Any]:
    from app.services.off_platform import ALL_KINDS
    inserted = updated = skipped = 0
    errors: list[str] = []
    for i, r in enumerate(rows, start=2):
        try:
            dt = _to_date(r.get("dt"))
            kind = _to_str(r.get("kind"))
            qty = _to_int(r.get("qty"))
            if not dt or not kind or not qty or qty <= 0:
                errors.append(f"строка {i}: нужны dt, kind, qty>0")
                skipped += 1
                continue
            if kind not in ALL_KINDS:
                errors.append(f"строка {i}: неизвестный kind {kind!r}")
                skipped += 1
                continue
            nm_id = _to_int(r.get("nm_id"))
            unit_cost = _to_decimal(r.get("unit_cost")) or Decimal(0)
            row_id = _to_int(r.get("id"))
            existing = (
                await session.get(OffPlatformStockMovement, row_id) if row_id else None
            )
            if nm_id is not None:
                await session.execute(
                    pg_insert(Product)
                    .values(tenant_id=get_tenant(session), nm_id=nm_id)
                    .on_conflict_do_nothing(index_elements=["nm_id"])
                )
            kwargs = dict(
                dt=dt, nm_id=nm_id, kind=kind, qty=int(qty),
                unit_cost=unit_cost, comment=_to_str(r.get("comment")),
            )
            if existing:
                for k, v in kwargs.items():
                    setattr(existing, k, v)
                updated += 1
            else:
                session.add(OffPlatformStockMovement(**kwargs))
                inserted += 1
        except Exception as e:
            errors.append(f"строка {i}: {e}")
            skipped += 1
    return {"inserted": inserted, "updated": updated, "skipped": skipped, "errors": errors}


async def _export_product_groups(session: AsyncSession) -> list[list[Any]]:
    rows = (await session.execute(
        select(ProductGroup).order_by(ProductGroup.name)
    )).scalars().all()
    return [[g.id, g.name, g.manager_name, g.color, g.comment] for g in rows]


async def _import_product_groups(
    session: AsyncSession, rows: list[dict[str, Any]]
) -> dict[str, Any]:
    inserted = updated = skipped = 0
    errors: list[str] = []
    for i, r in enumerate(rows, start=2):
        try:
            name = _to_str(r.get("name"))
            if not name:
                skipped += 1
                continue
            existing = (await session.execute(
                select(ProductGroup).where(ProductGroup.name == name)
            )).scalars().first()
            kwargs = dict(
                name=name,
                manager_name=_to_str(r.get("manager_name")),
                color=_to_str(r.get("color")),
                comment=_to_str(r.get("comment")),
            )
            if existing:
                for k, v in kwargs.items():
                    setattr(existing, k, v)
                updated += 1
            else:
                session.add(ProductGroup(**kwargs))
                inserted += 1
        except Exception as e:
            errors.append(f"строка {i}: {e}")
            skipped += 1
    return {"inserted": inserted, "updated": updated, "skipped": skipped, "errors": errors}


async def _export_pg_assignments(session: AsyncSession) -> list[list[Any]]:
    rows = (await session.execute(
        select(ProductGroup.name, ProductGroupAssignment.nm_id)
        .join(ProductGroupAssignment, ProductGroup.id == ProductGroupAssignment.group_id)
        .order_by(ProductGroup.name, ProductGroupAssignment.nm_id)
    )).all()
    return [[name, nm_id] for name, nm_id in rows]


async def _import_pg_assignments(
    session: AsyncSession, rows: list[dict[str, Any]]
) -> dict[str, Any]:
    inserted = updated = skipped = 0
    errors: list[str] = []
    # cache groups by name
    groups = (await session.execute(select(ProductGroup.id, ProductGroup.name))).all()
    by_name = {g.name: g.id for g in groups}
    for i, r in enumerate(rows, start=2):
        try:
            group_name = _to_str(r.get("group_name"))
            nm_id = _to_int(r.get("nm_id"))
            if not group_name or not nm_id:
                errors.append(f"строка {i}: нужны group_name и nm_id")
                skipped += 1
                continue
            gid = by_name.get(group_name)
            if not gid:
                errors.append(f"строка {i}: группа {group_name!r} не найдена")
                skipped += 1
                continue
            # ensure product exists for FK
            await session.execute(
                pg_insert(Product)
                .values(tenant_id=get_tenant(session), nm_id=nm_id)
                .on_conflict_do_nothing(index_elements=["nm_id"])
            )
            existing = (await session.execute(
                select(ProductGroupAssignment).where(
                    ProductGroupAssignment.group_id == gid,
                    ProductGroupAssignment.nm_id == nm_id,
                )
            )).scalars().first()
            if existing:
                skipped += 1
            else:
                session.add(ProductGroupAssignment(group_id=gid, nm_id=nm_id))
                inserted += 1
        except Exception as e:
            errors.append(f"строка {i}: {e}")
            skipped += 1
    return {"inserted": inserted, "updated": updated, "skipped": skipped, "errors": errors}


async def _export_supplies(session: AsyncSession) -> list[list[Any]]:
    rows = (await session.execute(select(Supply).order_by(Supply.supply_date.desc(), Supply.id.desc()))).scalars().all()
    return [
        [
            s.id,
            s.nm_id,
            s.vendor_code,
            s.supply_date,
            int(s.qty or 0),
            float(s.cost_per_unit) if s.cost_per_unit is not None else None,
            s.currency,
            s.vendor,
            s.invoice_number,
            s.paid_status,
            s.paid_date,
            float(s.paid_amount) if s.paid_amount is not None else None,
            s.notes,
        ]
        for s in rows
    ]


async def _import_supplies(
    session: AsyncSession, rows: list[dict[str, Any]]
) -> dict[str, Any]:
    """Upsert по id если задан, иначе insert. Натуральный ключ
    (nm_id+supply_date+invoice_number) — не enforce'им, потому что invoice
    бывает пустой. Если хотите редактировать существующую запись — экспортируйте
    сначала, чтобы получить id в первой колонке.
    """
    inserted = updated = skipped = 0
    errors: list[str] = []
    for i, raw in enumerate(rows, start=2):
        try:
            sid = _to_int(raw.get("id"))
            supply_date = _to_date(raw.get("supply_date"))
            if supply_date is None:
                raise ValueError("supply_date обязательна")
            qty = _to_int(raw.get("qty")) or 0
            if qty <= 0:
                raise ValueError("qty должна быть > 0")
            cost_per_unit = _to_decimal(raw.get("cost_per_unit")) or 0
            total_cost = cost_per_unit * qty
            paid_status = (raw.get("paid_status") or "unpaid").strip().lower()
            if paid_status not in ("unpaid", "partial", "paid"):
                paid_status = "unpaid"
            payload = {
                "nm_id": _to_int(raw.get("nm_id")),
                "vendor_code": _to_str(raw.get("vendor_code")),
                "supply_date": supply_date,
                "qty": qty,
                "cost_per_unit": cost_per_unit,
                "total_cost": total_cost,
                "currency": (raw.get("currency") or "RUB").strip().upper()[:8] or "RUB",
                "vendor": _to_str(raw.get("vendor")),
                "invoice_number": _to_str(raw.get("invoice_number")),
                "paid_status": paid_status,
                "paid_date": _to_date(raw.get("paid_date")),
                "paid_amount": _to_decimal(raw.get("paid_amount")),
                "notes": _to_str(raw.get("notes")),
            }
            if sid:
                existing = await session.get(Supply, sid)
                if existing is None:
                    # id указан но не найден — создаём новую запись (id выставит SEQ)
                    session.add(Supply(**payload))
                    inserted += 1
                else:
                    for k, v in payload.items():
                        setattr(existing, k, v)
                    updated += 1
            else:
                session.add(Supply(**payload))
                inserted += 1
        except Exception as e:
            errors.append(f"строка {i}: {e}")
            skipped += 1
    return {"inserted": inserted, "updated": updated, "skipped": skipped, "errors": errors}


async def _export_jam_queries(session: AsyncSession) -> list[list[Any]]:
    from app.db.models import JamQuery

    rows = (
        await session.execute(
            select(JamQuery).order_by(JamQuery.nm_id, JamQuery.period_start, JamQuery.id)
        )
    ).scalars().all()
    return [
        [r.nm_id, r.query, r.period_start, r.period_end, r.orders, r.clicks, r.views, r.ad_spent]
        for r in rows
    ]


async def _import_jam_queries(
    session: AsyncSession, rows: list[dict[str, Any]]
) -> dict[str, Any]:
    """Идемпотентный импорт. Уникальный ключ — (tenant, nm_id, query, period_start)."""
    from app.db.models import JamQuery
    from app.services.jam import upsert_jam_query

    inserted = updated = skipped = 0
    errors: list[str] = []
    for i, r in enumerate(rows, start=2):
        try:
            nm_id = _to_int(r.get("nm_id"))
            query = _to_str(r.get("query"))
            period_start = _to_date(r.get("period_start"))
            period_end = _to_date(r.get("period_end")) or period_start
            if not (nm_id and query and period_start):
                errors.append(f"строка {i}: nm_id, query, period_start обязательны")
                skipped += 1
                continue
            was_existing = (
                await session.execute(
                    select(JamQuery.id).where(
                        JamQuery.nm_id == nm_id,
                        JamQuery.query == query,
                        JamQuery.period_start == period_start,
                    )
                )
            ).scalar_one_or_none() is not None
            await upsert_jam_query(
                session,
                nm_id=int(nm_id),
                query=query,
                period_start=period_start,
                period_end=period_end,
                orders=_to_int(r.get("orders")) or 0,
                clicks=_to_int(r.get("clicks")) or 0,
                views=_to_int(r.get("views")) or 0,
                ad_spent=float(_to_decimal(r.get("ad_spent")) or 0),
            )
            if was_existing:
                updated += 1
            else:
                inserted += 1
        except Exception as e:
            errors.append(f"строка {i}: {e}")
            skipped += 1
    return {"inserted": inserted, "updated": updated, "skipped": skipped, "errors": errors}


REGISTRY: dict[str, tuple[ExportFn, ImportFn]] = {
    "products": (_export_products, _import_products),
    "cogs": (_export_cogs, _import_cogs),
    "opex_categories": (_export_opex_categories, _import_opex_categories),
    "opex_entries": (_export_opex_entries, _import_opex_entries),
    "artificial_orders": (_export_artif, _import_artif),
    "external_ad_costs": (_export_ext_ad, _import_ext_ad),
    "sales_plans": (_export_plans, _import_plans),
    "wb_tariff_categories": (_export_tariffs, _import_tariffs),
    "settings": (_export_settings, _import_settings),
    "setting_timeline": (_export_setting_timeline, _import_setting_timeline),
    "off_platform_stock": (_export_off_platform, _import_off_platform),
    "product_groups": (_export_product_groups, _import_product_groups),
    "product_group_assignments": (_export_pg_assignments, _import_pg_assignments),
    "supplies": (_export_supplies, _import_supplies),
    "jam_queries": (_export_jam_queries, _import_jam_queries),
}


def _excel_cell(v: Any) -> Any:
    if isinstance(v, Decimal):
        return float(v)
    return v


async def export_excel(session: AsyncSession, *, entity: str) -> bytes:
    if entity not in REGISTRY:
        raise ValueError(f"unknown entity: {entity}")
    headers = SCHEMAS[entity]
    rows = await REGISTRY[entity][0](session)
    wb = Workbook()
    ws = wb.active
    ws.title = entity[:31]
    ws.append(headers)
    for row in rows:
        ws.append([_excel_cell(v) for v in row])
    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = max(12, len(header) + 2)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def import_excel(
    session: AsyncSession, *, entity: str, file_bytes: bytes
) -> dict[str, Any]:
    if entity not in REGISTRY:
        raise ValueError(f"unknown entity: {entity}")
    expected = set(SCHEMAS[entity])
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    iter_ = ws.iter_rows(values_only=True)
    header_raw = next(iter_, None)
    if not header_raw:
        return {"inserted": 0, "updated": 0, "skipped": 0, "errors": ["пустой файл"]}
    header = [str(h).strip() if h is not None else "" for h in header_raw]
    rows: list[dict[str, Any]] = []
    for r in iter_:
        if r is None or all(c is None or c == "" for c in r):
            continue
        d: dict[str, Any] = {}
        for i, name in enumerate(header):
            if name and name in expected and i < len(r):
                d[name] = r[i]
        rows.append(d)
    res = await REGISTRY[entity][1](session, rows)
    await session.commit()
    return res
