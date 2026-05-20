"""UNIT-план — backend XLSX export 1:1 с эталонным Excel LeymanKids.

Структура листа «UNIT» воспроизводит эталон:

- **R1** — глобальные константы в фиксированных ячейках:
  E1='Обновлен:', F1=date.today(), J1=velocity_days, P1=wb_club_pct,
  R1=spp_default_pct, T1=wb_wallet_pct, V1=acquiring_pct,
  Z1='ИЛ', AA1=il_coef, AB1='ИРП', AC1=irp_coef,
  AN1=marketing_pct, AP1=tax_pct, AQ1=vat_mode_str, AR1=vat_pct,
  AS1=acceptance_rub_per_liter.

- **R2** — headers (русские названия 1:1).

- **R3+** — данные по SKU. Каждая строка = одна позиция из `compute_row`.

Дроби (`discount_pct`, `wb_club_pct`, `spp_pct`, `wb_wallet_pct`,
`commission_pct`, `acquiring_pct`, `commission_total_pct`, `buyout_pct`,
`logistics_share`, `storage_share`, `cogs_share`, `marketing_pct`,
`tax_pct`, `vat_pct`, `acceptance_share`, `margin_pct`, `roi_pct`,
`warehouse_coef_pct`) пишутся **как доли 0-1** с cell-format `'0.00%'` —
как в эталоне.

Рубли — числа с 2 знаками. None → пустая ячейка.
"""
from __future__ import annotations

import io
from datetime import date
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SHEET_NAME = "UNIT"

# ---------------------------------------------------------------------------
# Excel column layout (A..BF). Order matters — index → (column_letter, header,
# dto_field, kind).
#
# `kind`:
#   'str'     — текстовое значение
#   'int'     — целое (None → пусто)
#   'num'     — Decimal/float с 2 знаками (формат '0.00')
#   'pct'     — доля 0-1, формат '0.00%'
#   'bool_ru' — bool → 'Да'/'Нет'
#   'check'   — bool → 'OK'/'MISMATCH'
# ---------------------------------------------------------------------------
COLUMNS: list[tuple[str, str, str | None, str]] = [
    ("A",  "Склад",                                  "warehouse",            "str"),
    ("B",  "Литры",                                  "volume_l",             "num"),
    ("C",  "Бренд",                                  "brand",                "str"),
    ("D",  "Предмет",                                "subject",              "str"),
    ("E",  "Артикул продавца",                       "vendor_code",          "str"),
    ("F",  "Артикул WB",                             "nm_id",                "int"),
    ("G",  "Остаток на всех складах WB",             "stock_wb",             "int"),
    ("H",  "Остаток товара FBS",                     "stock_fbs",            "int"),
    ("I",  "Остаток с учетом  % выкупа",             "stock_effective",      "num"),
    ("J",  "закончится через, дн",                   "days_to_stockout",     "num"),
    ("K",  "Базовая цена",                           "base_price",           "num"),
    ("L",  "Скидка",                                 "discount_pct",         "pct"),
    ("M",  "Скидка с сайта ВБ",                      "discount_pct",         "pct"),
    ("N",  "Проверка",                               None,                   "check"),  # M == L
    ("O",  "Цена продажи (без СПП)",                 "price_after_discount", "num"),
    ("P",  "ВБ Клуб",                                "wb_club_pct",          "pct"),
    ("Q",  "Цена с ВБ клуб",                         "price_after_wb_club",  "num"),
    ("R",  "Размер СПП",                             "spp_pct",              "pct"),
    ("S",  "Цена с СПП",                             "price_after_spp",      "num"),
    ("T",  "Цена с WB кошелек",                      "price_final",          "num"),
    ("U",  "Комиссия,%",                             "commission_pct",       "pct"),
    ("V",  "Эквайринг",                              "acquiring_pct",        "pct"),
    ("W",  "Общая комиссия (вкл эквайринг)",         "commission_total_pct", "pct"),
    ("X",  "Комиссия,руб.",                          "commission_rub",       "num"),
    ("Y",  "FBS",                                    "is_fbs",               "bool_ru"),
    ("Z",  "Логистика, тариф",                       "logistics_box_rub",    "num"),
    ("AA", "Монопаллеты (Да/Нет)",                   "is_monopallet",        "bool_ru"),
    ("AB", "Сколько товара на монопаллете",          "items_per_monopallet", "int"),
    ("AC", "Логистика, тариф по монопаллетам",       "logistics_pallet_rub", "num"),
    ("AD", "% выкупа",                               "buyout_pct",           "pct"),
    ("AE", "Коэф-т склада, %",                       "warehouse_coef_pct",   "pct"),
    ("AF", "Логистика, в руб.",                      "logistics_rub",        "num"),
    ("AG", "Обратная логистика, в руб.",             "reverse_logistics_rub", "num"),
    ("AH", "Логистика, % от продажи",                "logistics_share",      "pct"),
    ("AI", "Хранение, в рублях",                     "storage_rub",          "num"),
    ("AJ", "Хранение, % от продаж",                  "storage_share",        "pct"),
    ("AK", "Себестоимость, в рублях",                "cogs_rub",             "num"),
    ("AL", "Себестоимость, % от продаж",             "cogs_share",           "pct"),
    ("AM", "Реклама в рублях",                       "marketing_rub",        "num"),
    ("AN", "Реклама, % от продаж",                   "marketing_pct",        "pct"),
    ("AO", "Налоги, в рублях ",                      "tax_rub",              "num"),
    ("AP", "Налоги, % от продажи",                   "tax_pct",              "pct"),
    ("AQ", "НДС, в рублях ",                         "vat_rub",              "num"),
    ("AR", "НДС, % от продажи",                      "vat_pct",              "pct"),
    ("AS", "Платная приемка, в рублях",              "acceptance_rub",       "num"),
    ("AT", "Платная приемка, % от продаж",           "acceptance_share",     "pct"),
    ("AU", "Прибыль в рублях",                       "profit_rub",           "num"),
    ("AV", "Маржинальность, прибыль % от продажи",   "margin_pct",           "pct"),
    ("AW", "Рентабельность, прибыль % от сс",        "roi_pct",              "pct"),
    ("AX", "Сезон",                                  "season_label",         "str"),
    ("AY", "пол",                                    "gender_label",         "str"),
    ("AZ", "АВС",                                    "abc_label",            "str"),
    # Historical snapshot columns (BA..BF) — pre-computed loader'ом из БД.
    # Передаются через HistoricalSnapshot → UnitPlanRowDTO (см. UNIT_PLAN.md §4).
    # Если значения нет (период не задан / 0 заказов) → пустая ячейка.
    ("BA", "Чистая прибыль  нед мая",                "profit_week_1",        "num"),
    ("BB", "Заказано период 1",                      "orders_period_1",      "int"),
    ("BC", "Выкуплено период 1",                     "sold_period_1",        "int"),
    ("BD", "Заказано период 2",                      "orders_period_2",      "int"),
    ("BE", "Заказано период 3",                      "orders_period_3",      "int"),
    ("BF", "Прогноз остатка",                        "stock_forecast",       "num"),
]


# ---------------------------------------------------------------------------
# Value coercion helpers
# ---------------------------------------------------------------------------


def _to_float(value: Any) -> float | None:
    """Decimal/str/int/float → float. None для пустых/нечисловых."""
    if value is None:
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, str):
        if value == "":
            return None
        try:
            return float(value)
        except ValueError:
            return None
    return None


def _coerce_cell(kind: str, value: Any) -> tuple[Any, str | None]:
    """Возвращает (cell_value, number_format)."""
    if value is None and kind not in ("check",):
        return None, None

    if kind == "str":
        return ("" if value is None else str(value)), None

    if kind == "int":
        v = _to_float(value)
        if v is None:
            return None, None
        return int(v), "0"

    if kind == "num":
        v = _to_float(value)
        if v is None:
            return None, None
        return v, "0.00"

    if kind == "pct":
        v = _to_float(value)
        if v is None:
            return None, None
        return v, "0.00%"

    if kind == "bool_ru":
        if value is None:
            return "Нет", None
        return ("Да" if bool(value) else "Нет"), None

    if kind == "check":
        # value передаётся явно (см. сборку строки)
        if value is None:
            return "", None
        return ("OK" if value else "MISMATCH"), None

    return value, None


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


_VAT_MODE_LABELS = {
    "include": "Да, включаем",
    "exclude": "Да, не включаем",
    "none": "не включаем",
}


def _r1_global_constants(ws, config: dict[str, Any]) -> None:
    """Записать R1 — глобальные константы в фиксированных ячейках."""
    today = date.today()

    def put(cell: str, value: Any, num_format: str | None = None) -> None:
        c = ws[cell]
        c.value = value
        if num_format:
            c.number_format = num_format

    put("E1", "Обновлен:   ")
    put("F1", today, "yyyy-mm-dd")

    velocity_days = config.get("velocity_days") or 30
    put("J1", int(velocity_days), "0")

    wb_club_pct = _to_float(config.get("wb_club_pct")) or 0.0
    put("P1", wb_club_pct, "0.00%")

    spp_default_pct = _to_float(config.get("spp_default_pct")) or 0.0
    put("R1", spp_default_pct, "0.00%")

    wb_wallet_pct = _to_float(config.get("wb_wallet_pct")) or 0.0
    put("T1", wb_wallet_pct, "0.00%")

    acquiring_pct = _to_float(config.get("acquiring_pct")) or 0.0
    put("V1", acquiring_pct, "0.00%")

    put("Z1", "ИЛ")
    il_coef = _to_float(config.get("il_coef")) or 1.16
    put("AA1", il_coef, "0.00")

    put("AB1", "ИРП")
    irp_coef = _to_float(config.get("irp_coef")) or 0.017
    put("AC1", irp_coef, "0.000")

    marketing_pct = _to_float(config.get("marketing_pct")) or 0.0
    put("AN1", marketing_pct, "0.00%")

    tax_pct = _to_float(config.get("tax_pct")) or 0.0
    put("AP1", tax_pct, "0.00%")

    vat_mode = config.get("vat_mode") or "none"
    put("AQ1", _VAT_MODE_LABELS.get(str(vat_mode), str(vat_mode)))

    vat_pct = _to_float(config.get("vat_pct")) or 0.0
    put("AR1", vat_pct, "0.00%")

    acceptance = _to_float(config.get("acceptance_rub_per_liter")) or 0.0
    put("AS1", acceptance, "0.00")

    # Стиль для R1 — мягкая заливка для визуального отделения от данных.
    fill = PatternFill("solid", fgColor="FFF2E5")
    bold = Font(bold=True)
    for col_letter in ("E", "Z", "AB"):
        ws[f"{col_letter}1"].font = bold
    for col, _hdr, _f, _k in COLUMNS:
        # Только подсветить cells где реально что-то лежит на R1
        if ws[f"{col}1"].value not in (None, ""):
            ws[f"{col}1"].fill = fill


def _write_headers(ws) -> None:
    """R2 — headers, bold + freeze."""
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    for col, header, _field, _kind in COLUMNS:
        c = ws[f"{col}2"]
        c.value = header
        c.font = bold
        c.fill = header_fill
        c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "G3"  # фиксируем первые 6 колонок и 2 верхние строки
    ws.row_dimensions[2].height = 36


def _write_row(ws, excel_row: int, row_data: dict[str, Any]) -> None:
    """Записать одну строку данных в R{excel_row}."""
    discount_value = row_data.get("discount_pct")
    discount_check_value = row_data.get("discount_pct")

    for col, _header, field, kind in COLUMNS:
        cell = ws[f"{col}{excel_row}"]
        if kind == "check":
            # N: проверка L == M. У нас обе берутся из discount_pct, всегда OK,
            # но если значение None — оставляем пусто.
            if discount_value is None:
                cell.value = ""
            else:
                cell.value = "OK"
            continue
        if field is None:
            cell.value = None
            continue
        raw = row_data.get(field)
        value, num_format = _coerce_cell(kind, raw)
        cell.value = value
        if num_format:
            cell.number_format = num_format
    # Чтобы не получить "0.0" для буквенных колонок:
    _ = discount_check_value  # placeholder для будущих сверок


def _set_column_widths(ws) -> None:
    """Базовые ширины колонок — читаемые для типовых значений."""
    widths = {
        "A": 16, "B": 8, "C": 16, "D": 22, "E": 18, "F": 12,
        "G": 10, "H": 10, "I": 12, "J": 12,
        "K": 12, "L": 10, "M": 10, "N": 10, "O": 12, "P": 8,
        "Q": 12, "R": 10, "S": 12, "T": 12,
        "U": 10, "V": 10, "W": 10, "X": 12, "Y": 8,
        "Z": 12, "AA": 10, "AB": 10, "AC": 12, "AD": 10, "AE": 10,
        "AF": 12, "AG": 12, "AH": 10,
        "AI": 12, "AJ": 10, "AK": 14, "AL": 10,
        "AM": 12, "AN": 10, "AO": 12, "AP": 10, "AQ": 12, "AR": 10,
        "AS": 12, "AT": 10, "AU": 14, "AV": 12, "AW": 12,
        "AX": 10, "AY": 6, "AZ": 6,
        "BA": 14, "BB": 12, "BC": 12, "BD": 12, "BE": 12, "BF": 14,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_unit_plan_xlsx(
    rows: list[dict[str, Any]],
    config: dict[str, Any],
    filters_applied: dict[str, Any] | None = None,
) -> bytes:
    """Сгенерировать .xlsx идентичный по структуре LeymanKids UNIT_план.

    Args:
        rows: список dict (асерт `UnitPlanRowDTO` сериализован через
              `dataclasses.asdict` + Decimal→str).
        config: глобальные константы (см. `_global_config_to_dict`
                из api/unit_plan.py). Допустимы str-Decimal, int, float.
        filters_applied: для имени файла / комментариев (не пишем в R1, чтобы
                         не плыли координаты констант).

    Returns:
        bytes — .xlsx содержимое.
    """
    _ = filters_applied  # пока не используем (имя файла формирует endpoint)

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    _r1_global_constants(ws, config)
    _write_headers(ws)

    for idx, row in enumerate(rows, start=3):
        _write_row(ws, idx, row)

    _set_column_widths(ws)

    # Гарантируем что любая «соседняя» дефолтная вкладка не утечёт.
    for other in list(wb.sheetnames):
        if other != SHEET_NAME:
            del wb[other]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


__all__ = ["build_unit_plan_xlsx", "COLUMNS", "SHEET_NAME"]
