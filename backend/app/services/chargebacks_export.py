"""XLSX-экспорт реестра претензий по chargebacks (LEAD-014).

PDF не делаем — у бухгалтера обычно Excel-workflow, и XLSX можно отредактировать
перед подачей в WB-поддержку. Если PDF понадобится — добавляется отдельно
через `reportlab` (новая зависимость).

Формат: один лист «Претензии», строки — chargebacks, столбцы по канонической
структуре. Заголовок с периодом и количеством, итоговая строка с суммами.
"""
from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.db.models import Chargeback
from app.services.chargebacks import CATEGORY_LABELS, STATUS_LABELS


HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TOTAL_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
TOTAL_FONT = Font(bold=True)


def build_chargebacks_xlsx(
    items: Iterable[Chargeback],
    *,
    period_label: str = "",
    tenant_name: str = "",
) -> bytes:
    """Собирает XLSX-реестр претензий. Возвращает bytes для FastAPI Response."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Претензии"

    # Шапка
    ws["A1"] = f"Реестр претензий{f' — {tenant_name}' if tenant_name else ''}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:H1")
    if period_label:
        ws["A2"] = f"Период: {period_label}"
        ws.merge_cells("A2:H2")
    ws["A3"] = f"Сгенерировано: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A3"].font = Font(italic=True, color="6B7280")
    ws.merge_cells("A3:H3")

    # Колонки
    headers = [
        "Дата операции",
        "Категория",
        "Тип WB",
        "SKU (nm_id)",
        "Сумма ₽",
        "Статус",
        "Текст претензии",
        "rrd_id",
    ]
    row_idx = 5
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Данные
    row_idx = 6
    total_sum = Decimal("0")
    count = 0
    for c in items:
        ws.cell(row=row_idx, column=1, value=c.operation_dt.isoformat() if c.operation_dt else "")
        ws.cell(
            row=row_idx,
            column=2,
            value=CATEGORY_LABELS.get(c.category, c.category),
        )
        ws.cell(row=row_idx, column=3, value=c.supplier_oper_name)
        ws.cell(row=row_idx, column=4, value=int(c.nm_id) if c.nm_id else "")
        amount_cell = ws.cell(row=row_idx, column=5, value=float(c.amount_rub))
        amount_cell.number_format = '#,##0.00'
        amount_cell.alignment = Alignment(horizontal="right")
        ws.cell(row=row_idx, column=6, value=STATUS_LABELS.get(c.status, c.status))
        # Текст претензии: либо из БД, либо из дефолтного шаблона (но шаблоны
        # будем подставлять на фронте, тут просто текст)
        ws.cell(row=row_idx, column=7, value=c.claim_text or "")
        ws.cell(row=row_idx, column=8, value=str(c.rrd_id))
        total_sum += c.amount_rub or Decimal("0")
        count += 1
        row_idx += 1

    # Итого
    total_row = row_idx
    ws.cell(row=total_row, column=1, value=f"Итого: {count} шт.")
    ws.cell(row=total_row, column=1).font = TOTAL_FONT
    ws.cell(row=total_row, column=1).fill = TOTAL_FILL
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    total_cell = ws.cell(row=total_row, column=5, value=float(total_sum))
    total_cell.number_format = '#,##0.00'
    total_cell.font = TOTAL_FONT
    total_cell.fill = TOTAL_FILL
    total_cell.alignment = Alignment(horizontal="right")
    for col in range(6, 9):
        ws.cell(row=total_row, column=col).fill = TOTAL_FILL

    # Ширины колонок
    widths = [13, 22, 28, 14, 13, 16, 50, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze pane на header row
    ws.freeze_panes = "A6"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
