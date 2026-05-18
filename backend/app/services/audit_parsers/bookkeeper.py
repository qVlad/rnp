"""Парсер XLSX от бухгалтера с настраиваемым mapping колонок.

Бухгалтеры используют разные форматы (1С / Контур / Моё дело / custom Excel).
В отличие от WB-парсера, мы НЕ знаем заранее какие колонки в файле — поэтому:

1. **Первая загрузка:** parse_bookkeeper(file_bytes, mapping=None) возвращает
   только список колонок и первые 5 строк превью → UI показывает wizard,
   юзер сопоставляет.
2. **Повторная загрузка** того же шаблона: parse_bookkeeper(file_bytes, mapping=...)
   возвращает нормализованный data_json.

Mapping JSON формат:
    {
        "column_to_code": {"Доходы от реализации": "revenue_gross", ...},
        "header_row": 1,                  -- строка с заголовками (1-based)
        "value_column": "сумма_руб",      -- если данные в одной "сумма" колонке
                                          -- (а строки разделены полем "статья")
        "sheet_name": "Отчёт",            -- если в файле несколько листов
    }

В v1 поддерживаем 2 формата:
- **"wide"** — каждая строка ОПиУ = отдельная колонка XLSX (mapping per column)
- **"long"** — одна колонка `статья` + одна `сумма` (mapping per row value)

`mapping.format` определяет какой.
"""
from __future__ import annotations

import io
from decimal import Decimal
from typing import Any

from openpyxl import load_workbook


def _norm(s: Any) -> str:
    return str(s or "").strip().lower()


def _to_decimal(v: Any) -> Decimal:
    if v is None or v == "":
        return Decimal("0")
    if isinstance(v, Decimal):
        return v
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    s = str(v).replace(",", ".").replace(" ", "").replace("\xa0", "")
    try:
        return Decimal(s)
    except Exception:
        return Decimal("0")


def preview_bookkeeper(file_bytes: bytes, *, max_rows: int = 5) -> dict[str, Any]:
    """Превью XLSX для UI mapping-wizard'а.

    Возвращает по каждому листу: имя, header (первая непустая строка) и
    первые N строк после header'а.
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    out: dict[str, Any] = {"sheets": []}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, max_row=max_rows + 5, values_only=True))
        # Найдём первую непустую строку как «вероятный header»
        header_idx = next(
            (i for i, r in enumerate(rows) if r and any(c is not None for c in r)),
            0,
        )
        header = [str(c or "") for c in (rows[header_idx] or [])]
        preview_rows = [
            [(str(c) if c is not None else "") for c in (r or [])]
            for r in rows[header_idx + 1 : header_idx + 1 + max_rows]
        ]
        out["sheets"].append(
            {
                "name": sheet_name,
                "suggested_header_row": header_idx + 1,
                "header": header,
                "preview_rows": preview_rows,
            }
        )
    return out


def _parse_wide(
    ws,
    *,
    header_row: int,
    column_to_code: dict[str, str],
) -> tuple[list[dict[str, Any]], int]:
    """«Wide» формат: каждая строка ОПиУ = отдельная колонка.

    Берём первую строку данных (header_row + 1). Если строк больше — складываем
    (на случай разбивки по месяцам внутри одного файла).
    """
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if header_row > len(rows):
        return [], 0
    headers = [str(h or "") for h in (rows[header_row - 1] or [])]
    # Индекс колонки → canonical code
    col_idx_to_code: dict[int, str] = {}
    for col_name, code in column_to_code.items():
        for i, h in enumerate(headers):
            if _norm(h) == _norm(col_name):
                col_idx_to_code[i] = code
                break

    sums: dict[str, Decimal] = {}
    rows_processed = 0
    for r in rows[header_row:]:
        if not r or all(c is None for c in r):
            continue
        rows_processed += 1
        for i, v in enumerate(r):
            if i in col_idx_to_code:
                sums[col_idx_to_code[i]] = sums.get(col_idx_to_code[i], Decimal("0")) + _to_decimal(v)

    return _sums_to_lines(sums), rows_processed


def _parse_long(
    ws,
    *,
    header_row: int,
    article_column: str,
    value_column: str,
    value_to_code: dict[str, str],
) -> tuple[list[dict[str, Any]], int]:
    """«Long» формат: одна колонка `статья`, одна `сумма`.

    Mapping: значение в article-колонке → canonical code.
    """
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    if header_row > len(rows):
        return [], 0
    headers = [_norm(h) for h in (rows[header_row - 1] or [])]
    try:
        art_idx = headers.index(_norm(article_column))
        val_idx = headers.index(_norm(value_column))
    except ValueError:
        return [], 0

    norm_value_to_code = {_norm(k): v for k, v in value_to_code.items()}
    sums: dict[str, Decimal] = {}
    rows_processed = 0
    for r in rows[header_row:]:
        if not r or len(r) <= max(art_idx, val_idx):
            continue
        article_val = _norm(r[art_idx])
        if not article_val:
            continue
        code = norm_value_to_code.get(article_val)
        if code is None:
            continue
        rows_processed += 1
        sums[code] = sums.get(code, Decimal("0")) + _to_decimal(r[val_idx])

    return _sums_to_lines(sums), rows_processed


# Лейблы для canonical-кодов (берём из audit_compare CANONICAL_LINES — но
# чтобы избежать циркулярного импорта, дублируем здесь).
_CANONICAL_LABELS: dict[str, str] = {
    "revenue_gross":   "Выручка (gross)",
    "revenue_returns": "Возвраты",
    "revenue_net":     "Чистая выручка",
    "commission_wb":   "Комиссия WB",
    "delivery_wb":     "Логистика WB",
    "storage_wb":      "Хранение WB",
    "acquiring":       "Эквайринг",
    "penalty":         "Штрафы",
    "deduction":       "Удержания",
    "ppvz_for_pay":    "К перечислению (ppvz_for_pay)",
    "ad_cost":         "Реклама",
    "cogs":            "Себестоимость",
    "vat_paid":        "НДС к уплате",
    "tax_paid":        "Налог (УСН/АУСН)",
    "net_profit":      "Чистая прибыль",
}


def _sums_to_lines(sums: dict[str, Decimal]) -> list[dict[str, Any]]:
    return [
        {
            "code": code,
            "label": _CANONICAL_LABELS.get(code, code),
            "amount": float(amount.quantize(Decimal("0.01"))),
        }
        for code, amount in sums.items()
    ]


def parse_bookkeeper(
    file_bytes: bytes,
    mapping: dict[str, Any] | None = None,
) -> dict[str, Any] | None:
    """Парсит bookkeeper XLSX согласно mapping.

    Если mapping=None — возвращает None (caller должен вызвать `preview_bookkeeper`
    для wizard'а).

    mapping format:
        {
            "format": "wide" | "long",
            "sheet_name": str | None,    # если None — берём первый
            "header_row": int,
            # для format="wide":
            "column_to_code": {col_name: canonical_code, ...}
            # для format="long":
            "article_column": str,
            "value_column": str,
            "value_to_code": {article_val: canonical_code, ...}
        }
    """
    if mapping is None:
        return None

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    sheet_name = mapping.get("sheet_name") or wb.sheetnames[0]
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    fmt = mapping.get("format", "wide")
    header_row = int(mapping.get("header_row") or 1)

    if fmt == "wide":
        lines, rows_processed = _parse_wide(
            ws,
            header_row=header_row,
            column_to_code=mapping.get("column_to_code") or {},
        )
    elif fmt == "long":
        lines, rows_processed = _parse_long(
            ws,
            header_row=header_row,
            article_column=mapping.get("article_column") or "",
            value_column=mapping.get("value_column") or "",
            value_to_code=mapping.get("value_to_code") or {},
        )
    else:
        raise ValueError(f"unknown bookkeeper format: {fmt!r}")

    return {
        "lines": lines,
        "raw_meta": {
            "sheet_name": sheet_name,
            "header_row": header_row,
            "rows_processed": rows_processed,
            "format": fmt,
        },
    }
