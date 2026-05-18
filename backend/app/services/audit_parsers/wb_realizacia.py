"""Парсер XLSX «Реализация» из WB-кабинета для аудит-режима.

Селлер скачивает в ЛК WB → Финансы → Финансовые отчёты → «Реализация в excel».
Формат — стандартный, WB меняет его редко, но изменения возможны. Парсер
устойчив: header-row ищется по характерным колонкам; если не найден — выбрасывает
понятную ошибку с показом первых строк (для UI fallback).

Маппит на canonical line_codes из `audit_compare.CANONICAL_LINES`.
"""
from __future__ import annotations

import io
import logging
from decimal import Decimal
from typing import Any

from openpyxl import load_workbook

log = logging.getLogger(__name__)


# Характерные заголовки колонок в WB XLSX (русские, иногда WB меняет регистр/пробелы).
# Ищем по подстроке case-insensitive — устойчиво к мелким вариациям WB.
WB_COLUMN_HINTS: dict[str, list[str]] = {
    # Тип документа: Продажа / Возврат / Корректировка
    "doc_type":          ["тип документа", "тип операции", "supplier_oper_name"],
    "retail_with_disc":  ["цена розничная с учётом согласованной скидки", "retail_price_withdisc_rub"],
    "ppvz_for_pay":      ["к перечислению продавцу", "ppvz_for_pay"],
    "commission":        ["размер кво за вычет", "размер кво", "wb комиссия"],
    "delivery":          ["услуги по доставке", "стоимость логистики"],
    "storage":           ["стоимость хранения", "услуги хранения"],
    "acquiring":         ["возмещение за выдачу", "эквайринг", "комиссия эквайер"],
    "penalty":           ["штрафы", "сумма штраф"],
    "deduction":         ["прочие удержания", "удержания"],
    "additional":        ["прочие доплаты", "доплата"],
}


class WbXlsxParseError(Exception):
    """Raised если структура файла не распознана."""

    def __init__(self, message: str, *, hints: list[str] | None = None) -> None:
        super().__init__(message)
        self.hints = hints or []


def _norm(s: Any) -> str:
    return str(s or "").strip().lower()


def _find_column_idx(headers: list[Any], hints: list[str]) -> int | None:
    """Возвращает 0-based index первой колонки чьё имя содержит любой из hints."""
    norm_headers = [_norm(h) for h in headers]
    for i, h in enumerate(norm_headers):
        for hint in hints:
            if hint.lower() in h:
                return i
    return None


def _find_header_row(ws) -> tuple[int, list[Any]]:
    """Сканирует первые 20 строк, ищет ту что содержит ≥3 ключевых колонки.

    WB иногда добавляет 1-5 строк с метаданными перед header'ом (период, селлер id).
    """
    rows = list(ws.iter_rows(min_row=1, max_row=20, values_only=True))
    required_groups = ["doc_type", "retail_with_disc", "ppvz_for_pay"]
    for row_idx, row in enumerate(rows):
        headers = list(row)
        if not any(headers):
            continue
        match_count = sum(
            1 for g in required_groups
            if _find_column_idx(headers, WB_COLUMN_HINTS[g]) is not None
        )
        if match_count >= 2:  # min 2 из 3 — допускаем мелкие отклонения
            return row_idx + 1, headers  # openpyxl rows 1-based
    raise WbXlsxParseError(
        "Не нашёл заголовок столбцов в первых 20 строках. "
        "Это файл WB-«Реализация»? Должны быть колонки «Тип документа», "
        "«Цена розничная», «К перечислению продавцу».",
        hints=[
            f"Row {i+1}: " + " | ".join(str(c or "")[:30] for c in (r or [])[:8])
            for i, r in enumerate(rows[:5])
        ],
    )


def _to_decimal(v: Any) -> Decimal:
    if v is None or v == "":
        return Decimal("0")
    if isinstance(v, Decimal):
        return v
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    s = str(v).replace(",", ".").replace(" ", "").replace("\xa0", "")
    try:
        return Decimal(s)
    except Exception:
        return Decimal("0")


def parse_wb_realizacia(file_bytes: bytes) -> dict[str, Any]:
    """Парсит WB XLSX «Реализация» и возвращает нормализованный data_json.

    Возвращает:
        {
            "lines": [
                {"code": "revenue_gross",   "label": "...", "amount": <float>},
                {"code": "revenue_returns", "label": "...", "amount": <float>},
                ...
            ],
            "raw_meta": {
                "sheet_name": "...",
                "header_row": int,
                "rows_processed": int,
                "doc_type_breakdown": {"Продажа": N, "Возврат": M, ...},
            }
        }

    Raises WbXlsxParseError если структура не распознана.
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    sheets = wb.sheetnames
    # WB обычно даёт один основной лист. Берём первый или с именем содержащим
    # «реализаци» (несколько вариаций спрашивает WB — Реализация / Реализации).
    target_sheet = next(
        (s for s in sheets if "реализаци" in s.lower()),
        sheets[0],
    )
    ws = wb[target_sheet]

    header_row, headers = _find_header_row(ws)

    # Собираем индексы интересующих нас колонок
    col_idx: dict[str, int | None] = {
        key: _find_column_idx(headers, hints)
        for key, hints in WB_COLUMN_HINTS.items()
    }

    # Aggregates по типу документа
    aggregates: dict[str, dict[str, Decimal]] = {}
    rows_processed = 0
    doc_type_counts: dict[str, int] = {}

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(row):
            continue
        doc_type_raw = (
            str(row[col_idx["doc_type"]] or "").strip()
            if col_idx["doc_type"] is not None
            else ""
        )
        if not doc_type_raw:
            continue
        rows_processed += 1
        doc_type_counts[doc_type_raw] = doc_type_counts.get(doc_type_raw, 0) + 1
        agg = aggregates.setdefault(
            doc_type_raw,
            {
                "retail_with_disc": Decimal("0"),
                "ppvz_for_pay":     Decimal("0"),
                "delivery":         Decimal("0"),
                "storage":          Decimal("0"),
                "acquiring":        Decimal("0"),
                "penalty":          Decimal("0"),
                "deduction":        Decimal("0"),
                "additional":       Decimal("0"),
            },
        )
        for key in agg:
            ix = col_idx.get(key)
            if ix is None:
                continue
            agg[key] += _to_decimal(row[ix])

    # Маппинг агрегатов в canonical line_codes
    sales = aggregates.get("Продажа", {})
    returns = aggregates.get("Возврат", {})

    def total(field: str) -> Decimal:
        return sum(
            (a.get(field, Decimal("0")) for a in aggregates.values()),
            Decimal("0"),
        )

    revenue_gross = sales.get("retail_with_disc", Decimal("0"))
    revenue_returns = returns.get("retail_with_disc", Decimal("0"))
    revenue_net = revenue_gross - revenue_returns
    # Комиссия = (retail − ppvz − acquiring) на Продажах, минус то же на Возвратах
    s_commission = (
        sales.get("retail_with_disc", Decimal("0"))
        - sales.get("ppvz_for_pay", Decimal("0"))
        - sales.get("acquiring", Decimal("0"))
    )
    r_commission = (
        returns.get("retail_with_disc", Decimal("0"))
        - returns.get("ppvz_for_pay", Decimal("0"))
        - returns.get("acquiring", Decimal("0"))
    )
    commission = s_commission - r_commission
    acquiring_net = sales.get("acquiring", Decimal("0")) - returns.get("acquiring", Decimal("0"))
    ppvz = sales.get("ppvz_for_pay", Decimal("0")) - returns.get("ppvz_for_pay", Decimal("0"))

    lines = [
        ("revenue_gross",   "Выручка (gross)",                revenue_gross),
        ("revenue_returns", "Возвраты",                       revenue_returns),
        ("revenue_net",     "Чистая выручка",                 revenue_net),
        ("commission_wb",   "Комиссия WB",                    commission),
        ("delivery_wb",     "Логистика WB",                   total("delivery")),
        ("storage_wb",      "Хранение WB",                    total("storage")),
        ("acquiring",       "Эквайринг",                      acquiring_net),
        ("penalty",         "Штрафы",                         total("penalty")),
        ("deduction",       "Удержания",                      total("deduction")),
        ("ppvz_for_pay",    "К перечислению (ppvz_for_pay)",  ppvz),
    ]

    return {
        "lines": [
            {"code": c, "label": l, "amount": float(a.quantize(Decimal("0.01")))}
            for c, l, a in lines
        ],
        "raw_meta": {
            "sheet_name": target_sheet,
            "header_row": header_row,
            "rows_processed": rows_processed,
            "doc_type_breakdown": doc_type_counts,
        },
    }
