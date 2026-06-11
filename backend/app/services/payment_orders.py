"""Импорт XLSX «История платежей» из ЛК WB.

Публичного API WB для этой страницы нет (см. wb-api-specialist research,
май 2026), используется ручной XLSX-импорт.

Ожидаемый формат файла:
    Лист с заголовком в первой строке. Колонки (порядок не важен,
    парсер находит по подстроке в имени):

      • «ID заявки на оплату»  → payment_order_id  (string, "4400004/53")
      • «Сумма»                → amount            (number, ₽)
      • «Валюта»               → currency          (string, "руб." → RUB)
      • «Дата создания»        → created_dt        (date, DD.MM.YYYY)
      • «Статус оплаты»        → status + paid_dt
                                  raw text:
                                  - "Оплата обрабатывается"
                                    → status='processing', paid_dt=NULL
                                  - "Оплата успешно проведена банком DD.MM.YYYY"
                                    → status='paid', paid_dt=<parsed>
                                  - всё остальное → status='unknown'
      • «Комментарий банка»    → bank_comment      (string)

Все строки upsert'ятся по PK (tenant_id, payment_order_id). Повторный
импорт того же файла идемпотентен. Изменения статуса (processing → paid)
обновляются.
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.models import WbPaymentOrder


# ── Маппинг колонок: ключи — наши поля, значения — список substring'ов
# для поиска в заголовке. Lowercase + casefold для устойчивости.
_COLUMN_MAP: dict[str, list[str]] = {
    "payment_order_id": ["id заявки", "id платежа", "номер заявки"],
    "amount": ["сумма"],
    "currency": ["валюта"],
    "created_dt": ["дата создания"],
    "status_raw": ["статус оплаты", "статус"],
    "bank_comment": ["комментарий банка", "комментарий"],
}

# Дата в статусе: "Оплата успешно проведена банком 13.05.2026"
_STATUS_DATE_RE = re.compile(r"(\d{2})\.(\d{2})\.(\d{4})")
_PROCESSING_RE = re.compile(r"обрабатыва", re.IGNORECASE)
_PAID_RE = re.compile(r"успешно|провед", re.IGNORECASE)
_FAILED_RE = re.compile(r"отказ|ошиб|отклон", re.IGNORECASE)


@dataclass
class ImportResult:
    rows_total: int = 0
    rows_inserted: int = 0
    rows_updated: int = 0
    rows_skipped: int = 0
    errors: list[str] | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "rows_total": self.rows_total,
            "rows_inserted": self.rows_inserted,
            "rows_updated": self.rows_updated,
            "rows_skipped": self.rows_skipped,
            "errors": self.errors or [],
        }


def _norm_header(s: Any) -> str:
    return str(s or "").strip().casefold()


def _match_column(header: str, candidates: list[str]) -> bool:
    h = _norm_header(header)
    return any(c.casefold() in h for c in candidates)


def _find_columns(header_row: list[Any]) -> dict[str, int]:
    """Возвращает {field_name: column_index} (0-based) для найденных полей."""
    out: dict[str, int] = {}
    for col_idx, cell in enumerate(header_row):
        for field, candidates in _COLUMN_MAP.items():
            if field in out:
                continue
            if _match_column(cell, candidates):
                out[field] = col_idx
                break
    return out


def _parse_date_ru(value: Any) -> date | None:
    """Парсит дату DD.MM.YYYY (или datetime/date как есть)."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    m = _STATUS_DATE_RE.search(s)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            return None
    # ISO fallback
    try:
        return datetime.fromisoformat(s.replace(" ", "T")).date()
    except (ValueError, TypeError):
        return None


def _parse_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None
    s = str(value).strip().replace(" ", "").replace(" ", "").replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _normalize_status(raw: str) -> tuple[str, date | None]:
    """Возвращает (нормализованный статус, дата зачисления если paid)."""
    if not raw:
        return ("unknown", None)
    if _PROCESSING_RE.search(raw):
        return ("processing", None)
    if _FAILED_RE.search(raw):
        return ("failed", None)
    if _PAID_RE.search(raw):
        return ("paid", _parse_date_ru(raw))
    return ("unknown", None)


def _normalize_currency(raw: Any) -> str:
    if not raw:
        return "RUB"
    s = str(raw).strip().lower()
    if "руб" in s or s == "rub" or s == "₽":
        return "RUB"
    if "usd" in s or "$" in s:
        return "USD"
    if "eur" in s or "€" in s:
        return "EUR"
    return s.upper()[:8]




def parse_payment_history_xlsx(
    file_bytes: bytes,
) -> tuple[list[dict[str, Any]], list[str]]:
    """Парсит XLSX. Возвращает (rows, errors).

    Каждая строка — dict со всеми полями WbPaymentOrder (без tenant_id).
    Ошибки — список человеческих сообщений (для показа юзеру).
    """
    errors: list[str] = []
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:
        return [], [f"Не удалось открыть XLSX: {e}"]

    # Берём первый лист
    ws = wb.active
    if ws is None:
        return [], ["В файле нет листов"]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return [], ["Файл пустой"]

    cols = _find_columns(list(header))
    required = ["payment_order_id", "amount", "created_dt", "status_raw"]
    missing = [r for r in required if r not in cols]
    if missing:
        errors.append(
            "В заголовке не найдены обязательные колонки: "
            + ", ".join(missing)
            + ". Ожидаются (порядок неважен): «ID заявки на оплату», «Сумма»,"
            " «Валюта», «Дата создания», «Статус оплаты», «Комментарий банка»."
        )
        return [], errors

    out: list[dict[str, Any]] = []
    for row_num, row in enumerate(rows_iter, start=2):
        if not row or all(c is None or c == "" for c in row):
            continue
        try:
            poid = str(row[cols["payment_order_id"]] or "").strip()
            if not poid:
                continue
            amount = _parse_decimal(row[cols["amount"]])
            created_dt = _parse_date_ru(row[cols["created_dt"]])
            status_raw = str(row[cols["status_raw"]] or "").strip()
            if amount is None or created_dt is None:
                errors.append(
                    f"Строка {row_num}: пропущена (пустая Сумма или Дата создания)"
                )
                continue
            status, paid_dt = _normalize_status(status_raw)
            currency = _normalize_currency(
                row[cols["currency"]] if "currency" in cols else None
            )
            bank_comment = (
                str(row[cols["bank_comment"]] or "").strip()
                if "bank_comment" in cols
                else None
            ) or None
            out.append({
                "payment_order_id": poid[:64],
                "created_dt": created_dt,
                "paid_dt": paid_dt,
                "amount": amount,
                "currency": currency,
                "status": status,
                "status_raw": status_raw[:255] or None,
                "bank_comment": (bank_comment or "")[:512] or None,
            })
        except Exception as e:
            errors.append(f"Строка {row_num}: ошибка парсинга — {e}")

    return out, errors


async def upsert_payment_orders(
    session: AsyncSession,
    rows: list[dict[str, Any]],
) -> ImportResult:
    """Upsert'ит строки в wb_payment_order. Tenant_id ставится через
    `_stamp_tenant`-механику на уровне session.info (как в sync/tasks.py)."""
    result = ImportResult(rows_total=len(rows))
    if not rows:
        return result

    # Считаем сколько уже было — для статистики inserted vs updated
    poids = [r["payment_order_id"] for r in rows]
    existing = (
        await session.execute(
            select(WbPaymentOrder.payment_order_id).where(
                WbPaymentOrder.payment_order_id.in_(poids)
            )
        )
    ).scalars().all()
    existing_set = set(existing)

    # tenant_id берём из session.info (set_tenant был вызван в depends)
    tid = session.sync_session.info.get("tenant_id")
    if tid is None:
        result.errors = ["tenant_id не задан в сессии (auth bug)"]
        return result

    values = [{**r, "tenant_id": tid} for r in rows]
    stmt = pg_insert(WbPaymentOrder).values(values)
    # Включаем все поля что есть в values; используем COALESCE-like
    # подход — обновляем поле только если оно есть в payload (для
    # Стас-формата upd_delivery_amount/period_end/report_type/buyout_returns_amount
    # передаются явно; для legacy «История платежей» — этих ключей нет).
    update_set: dict[str, Any] = {
        "created_dt": stmt.excluded.created_dt,
        "paid_dt": stmt.excluded.paid_dt,
        "amount": stmt.excluded.amount,
        "currency": stmt.excluded.currency,
        "status": stmt.excluded.status,
        "status_raw": stmt.excluded.status_raw,
        "bank_comment": stmt.excluded.bank_comment,
    }
    # Опциональные Стас-поля — обновляем только если хотя бы в одной
    # строке payload они есть (определяем по наличию ключа).
    if any("period_end" in r for r in rows):
        update_set["period_end"] = stmt.excluded.period_end
    if any("report_type" in r for r in rows):
        update_set["report_type"] = stmt.excluded.report_type
    if any("upd_delivery_amount" in r for r in rows):
        update_set["upd_delivery_amount"] = stmt.excluded.upd_delivery_amount
    if any("buyout_returns_amount" in r for r in rows):
        update_set["buyout_returns_amount"] = stmt.excluded.buyout_returns_amount
    stmt = stmt.on_conflict_do_update(
        index_elements=["tenant_id", "payment_order_id"],
        set_=update_set,
    )
    await session.execute(stmt)

    result.rows_inserted = sum(1 for r in rows if r["payment_order_id"] not in existing_set)
    result.rows_updated = sum(1 for r in rows if r["payment_order_id"] in existing_set)
    return result


# ── Стас «Разметка банка» — лист «Отчеты+УПД» ────────────────────────────────
# Источник истины для АУСН (TASK-DEV-068): Банк = «Итого к оплате» по «Дата
# оплаты», ВЗЗ/УПД по «Дата конца» (period_end). Воспроизводит лист «Итоги»
# копейка-в-копейку (проверено на апреле/мае 2026).
_STAS_COLUMN_MAP: dict[str, list[str]] = {
    "payment_order_id": ["№ отчета", "номер отчета", "n отчета"],
    "period_end": ["дата конца"],
    "created_dt": ["дата формирования"],
    "report_type": ["тип отчета"],
    "amount": ["итого к оплате"],
    "paid_dt": ["дата оплаты"],
    "upd_delivery_amount": ["упд доставка"],  # «УПД Доставка по выкупу»
    "buyout_returns_amount": ["возвраты выкупы"],
}


def _is_stas_sheet(header_row: list[Any]) -> bool:
    """Лист «Отчеты+УПД» определяем по наличию «Дата оплаты» + «УПД Доставка»."""
    cols = _find_columns_map(list(header_row), _STAS_COLUMN_MAP)
    return "paid_dt" in cols and "upd_delivery_amount" in cols and "amount" in cols


def _find_columns_map(header_row: list[Any], cmap: dict[str, list[str]]) -> dict[str, int]:
    out: dict[str, int] = {}
    for col_idx, cell in enumerate(header_row):
        for field, candidates in cmap.items():
            if field in out:
                continue
            if _match_column(cell, candidates):
                out[field] = col_idx
                break
    return out


def parse_stas_razmetka_xlsx(
    file_bytes: bytes,
) -> tuple[list[dict[str, Any]], list[str]]:
    """Парсит файл бухгалтера «Стас Разметка банка» (лист «Отчеты+УПД»).

    Возвращает (rows, errors) для upsert в wb_payment_order:
      payment_order_id = «№ отчета»; amount = «Итого к оплате» (→ Банк по
      «Дата оплаты»); paid_dt = «Дата оплаты» (есть → status=paid);
      period_end = «Дата конца»; report_type = «Тип отчета»;
      upd_delivery_amount = «УПД Доставка по выкупу»; buyout_returns_amount =
      «Возвраты выкупы». Все строки — единый набор ключей (для pg_insert).
    """
    errors: list[str] = []
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:
        return [], [f"Не удалось открыть XLSX: {e}"]

    # Ищем лист с заголовком «Дата оплаты» + «УПД Доставка».
    target_ws = None
    header: list[Any] = []
    for ws in wb.worksheets:
        try:
            first = next(ws.iter_rows(values_only=True))
        except StopIteration:
            continue
        if first and _is_stas_sheet(list(first)):
            target_ws = ws
            header = list(first)
            break
    if target_ws is None:
        return [], ["Лист «Отчеты+УПД» (с колонками «Дата оплаты» и «УПД Доставка по выкупу») не найден"]

    cols = _find_columns_map(header, _STAS_COLUMN_MAP)
    if "payment_order_id" not in cols or "amount" not in cols:
        return [], ["В листе нет обязательных колонок «№ отчета» / «Итого к оплате»"]

    rows_iter = target_ws.iter_rows(values_only=True)
    next(rows_iter)  # skip header
    out: list[dict[str, Any]] = []
    for row_num, row in enumerate(rows_iter, start=2):
        if not row or all(c is None or c == "" for c in row):
            continue
        try:
            poid = str(row[cols["payment_order_id"]] or "").strip()
            if not poid or not poid.replace(".0", "").isdigit():
                continue  # пропускаем итоговые/пустые строки
            poid = poid.replace(".0", "")
            amount = _parse_decimal(row[cols["amount"]]) or Decimal("0")
            paid_dt = _parse_date_ru(row[cols["paid_dt"]]) if "paid_dt" in cols else None
            period_end = _parse_date_ru(row[cols["period_end"]]) if "period_end" in cols else None
            created_dt = (
                _parse_date_ru(row[cols["created_dt"]]) if "created_dt" in cols else None
            ) or period_end
            report_type = (
                str(row[cols["report_type"]] or "").strip()[:32] or None
                if "report_type" in cols else None
            )
            upd = (
                _parse_decimal(row[cols["upd_delivery_amount"]])
                if "upd_delivery_amount" in cols else None
            ) or Decimal("0")
            buyout = (
                _parse_decimal(row[cols["buyout_returns_amount"]])
                if "buyout_returns_amount" in cols else None
            ) or Decimal("0")
            out.append({
                "payment_order_id": poid[:64],
                "created_dt": created_dt,
                "paid_dt": paid_dt,
                "amount": amount,
                "currency": "RUB",
                "status": "paid" if paid_dt is not None else "processing",
                "status_raw": None,
                "bank_comment": None,
                "period_end": period_end,
                "report_type": report_type,
                "upd_delivery_amount": upd,
                "buyout_returns_amount": buyout,
            })
        except Exception as e:
            errors.append(f"Строка {row_num}: ошибка парсинга — {e}")
    return out, errors


def parse_payment_xlsx_auto(
    file_bytes: bytes,
) -> tuple[list[dict[str, Any]], list[str], str]:
    """Авто-детект формата: «Стас Разметка» (лист «Отчеты+УПД») vs WB «История
    платежей». Возврат: (rows, errors, format_label)."""
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:
        return [], [f"Не удалось открыть XLSX: {e}"], "unknown"
    for ws in wb.worksheets:
        try:
            first = next(ws.iter_rows(values_only=True))
        except StopIteration:
            continue
        if first and _is_stas_sheet(list(first)):
            rows, errs = parse_stas_razmetka_xlsx(file_bytes)
            return rows, errs, "stas_razmetka"
    rows, errs = parse_payment_history_xlsx(file_bytes)
    return rows, errs, "payment_history"
