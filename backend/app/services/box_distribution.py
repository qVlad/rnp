"""Box Distribution — парсер файла «Распределение» + сборка выходного shk-excel.

DEV-091. Исходный Excel: листы брендов (Ink/Ld/Lk) со строками
`ШК короба | артикул продавца | Баркод | Размер | количество | Склад`.
Выход: shk-excel `Баркод товара | Кол-во товаров | ШК короба | Срок годности`.
"""
from __future__ import annotations

import io
from typing import Any

import openpyxl
from openpyxl import Workbook


# Безопасная нормализация складов (регистр/опечатки/разговорные). НЕ сливаем
# реально отдельные WB-склады (Шушары, Екатеринбург-Перспективная 14/17) — это
# делает пользователь через редактируемую карту алиасов.
WAREHOUSE_ALIASES_SEED: dict[str, str] = {
    "тула": "Тула",
    "элетросталь": "Электросталь",
    "невиномысск": "Невинномысск",
    # СПб-регион: все варианты в один склад (по решению пользователя 2026-06-20).
    "спб": "СПб",
    "питер": "СПб",
    "шушары": "СПб",
    # Екатеринбург-Перспективная 14 и 17 — в один склад.
    "екатеринбург-перспективная 14": "Екатеринбург-Перспективная",
    "екатеринбург-перспективная 17": "Екатеринбург-Перспективная",
}


def normalize_barcode(raw: Any) -> str:
    """`2049302632159.0` / `2049302632159 ` → `2049302632159`."""
    if raw is None:
        return ""
    s = str(raw).strip()
    if not s:
        return ""
    try:
        # числовой баркод (в т.ч. с хвостом .0 от float) → целое
        return str(int(float(s)))
    except (TypeError, ValueError):
        return s


def normalize_qty(raw: Any) -> int:
    if raw is None:
        return 0
    try:
        return int(round(float(str(raw).replace(",", ".").replace(" ", ""))))
    except (TypeError, ValueError):
        return 0


def normalize_warehouse(raw: Any, alias_map: dict[str, str] | None = None) -> str:
    """Канонизация склада: trim → seed-карта → пользовательская карта.

    Сопоставление по нижнему регистру; результат — каноническое имя.
    """
    if raw is None:
        return ""
    name = str(raw).strip()
    if not name:
        return ""
    key = name.lower()
    # seed (регистр/опечатки)
    canon = WAREHOUSE_ALIASES_SEED.get(key, name)
    # пользовательские слияния (ключи могут быть в любом регистре)
    if alias_map:
        lower_map = {str(k).strip().lower(): v for k, v in alias_map.items()}
        canon = lower_map.get(key, lower_map.get(canon.lower(), canon))
    return canon


def _find_header(rows: list[tuple], max_scan: int = 5) -> tuple[int, dict[str, int]] | None:
    """Найти строку-заголовок (в первых max_scan строках есть «короб») и
    смаппить нужные колонки по substring. Возвращает (row_index, {key: col})."""
    for i, r in enumerate(rows[:max_scan]):
        if not r:
            continue
        joined = " ".join(str(c).lower() for c in r if c is not None)
        if "короб" not in joined and "box code" not in joined:
            continue
        cols: dict[str, int] = {}
        for j, c in enumerate(r):
            s = str(c).lower() if c is not None else ""
            if not s:
                continue
            if "box" in cols and "wh" in cols and "bc" in cols and "qty" in cols:
                break
            if "короб" in s or "box code" in s:
                cols.setdefault("box", j)
            elif "артикул" in s:
                cols.setdefault("article", j)
            elif "баркод" in s:
                cols.setdefault("bc", j)
            elif "размер" in s:
                cols.setdefault("size", j)
            elif "кол" in s:  # количество / количество, шт
                cols.setdefault("qty", j)
            elif "склад" in s:
                cols.setdefault("wh", j)
        if "box" in cols and "bc" in cols and "wh" in cols and "qty" in cols:
            return i, cols
    return None


def parse_distribution_file(
    content: bytes, alias_map: dict[str, str] | None = None
) -> dict[str, Any]:
    """Распарсить все листы с заголовком короба. Бренд = имя листа.

    Возвращает {"rows": [...], "sheets": [...], "skipped": [...]}.
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    out_rows: list[dict[str, Any]] = []
    sheets_used: list[str] = []
    skipped: list[str] = []

    def cell(r: tuple, idx: int | None) -> Any:
        return r[idx] if idx is not None and idx < len(r) else None

    for ws in wb.worksheets:
        # Исключаем сводные/pivot-листы (дублируют данные брендов): «Свод»,
        # «Сводная таблица …». Парсим только листы-бренды (Ink/Ld/Lk и т.п.).
        if "свод" in ws.title.strip().lower():
            skipped.append(ws.title)
            continue
        rows = list(ws.iter_rows(values_only=True))
        hdr = _find_header(rows)
        if hdr is None:
            skipped.append(ws.title)
            continue
        hi, cols = hdr
        sheets_used.append(ws.title)
        for r in rows[hi + 1 :]:
            if not r or all(c is None for c in r):
                continue
            box = cell(r, cols.get("box"))
            barcode = normalize_barcode(cell(r, cols.get("bc")))
            wh_raw = cell(r, cols.get("wh"))
            wh = normalize_warehouse(wh_raw, alias_map)
            qty = normalize_qty(cell(r, cols.get("qty")))
            if not box or not barcode or not wh:
                continue
            out_rows.append(
                {
                    "brand": ws.title,
                    "src_box_code": str(box).strip(),
                    "vendor_article": (
                        str(cell(r, cols.get("article"))).strip()
                        if cell(r, cols.get("article")) is not None
                        else None
                    ),
                    "barcode": barcode,
                    "size": (
                        str(cell(r, cols.get("size"))).strip()
                        if cell(r, cols.get("size")) is not None
                        else None
                    ),
                    "qty": qty,
                    "warehouse": wh,
                    "warehouse_raw": (
                        str(wh_raw).strip() if wh_raw is not None else None
                    ),
                }
            )
    wb.close()
    return {"rows": out_rows, "sheets": sheets_used, "skipped": skipped}


def build_shk_xlsx(items: list[dict[str, Any]]) -> bytes:
    """items: [{barcode, qty, wb_box_code, expiry?}] → xlsx по шаблону page-excel.

    Колонки: Баркод товара | Кол-во товаров | ШК короба | Товар с кизом | Срок годности.
    «Товар с кизом» — всегда «да» (по требованию пользователя); срок годности — пусто.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(
        [
            "Баркод товара",
            "Кол-во товаров",
            "ШК короба",
            "Товар с кизом",
            "Срок годности",
        ]
    )
    for it in items:
        ws.append(
            [
                str(it.get("barcode") or ""),
                int(it.get("qty") or 0),
                str(it.get("wb_box_code") or ""),
                "да",
                it.get("expiry") or "",
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
