"""Выгрузки xlsx для WMS (TASK-DEV-098).

Ключевое: `build_state_xlsx` пишет ТОТ ЖЕ формат B, который читает
`packing_list.parse_receive_file`. Значит работает round-trip «выгрузил →
поправил в Excel → загрузил обратно», и одна и та же шапка описывает и
приёмку, и текущее состояние склада.
"""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# Шапка формата B. Первые две колонки — надстройка над PackingList; дальше
# идут колонки самого PackingList, чтобы файл поставщика грузился как есть.
STATE_HEADER = [
    "Склад",
    "Код ячейки",
    "No",
    "Box Code",
    "Barcode",
    "Size",
    "Qty",
    "Артикул",
    "nmID",
    "Название",
    "Бренд",
    "Статус",
    "Поставка",
]

CELLS_HEADER = ["Склад", "Код ячейки", "Зона", "Стеллаж", "Ярус", "Позиция", "Активна", "Примечание"]


def _autosize(ws: Any, header: list[str]) -> None:
    for i, name in enumerate(header, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(name) + 4)


def _style_header(ws: Any) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def build_state_xlsx(rows: list[dict[str, Any]]) -> bytes:
    """Текущее состояние склада в формате B (round-trip с приёмкой).

    `rows`: строки наличия из `stock.search`/`stock.stock` вида
    ``{warehouse_name, cell_code, src_no, box_code, barcode, size, qty,
    vendor_code, nm_id, name, brand, status_label, supply_ref}``.

    Как в PackingList, повторяющиеся значения короба печатаются только в его
    первой строке — файл читается человеком так же, как исходный от поставщика.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Склад"
    ws.append(STATE_HEADER)
    _style_header(ws)

    last_box: tuple[Any, Any] | None = None
    for row in rows:
        box_key = (row.get("warehouse_name"), row.get("box_code"))
        first_of_box = box_key != last_box
        last_box = box_key
        ws.append(
            [
                row.get("warehouse_name") or "",
                (row.get("cell_code") or "") if first_of_box else "",
                (row.get("src_no") or "") if first_of_box else "",
                row.get("box_code") or "",
                str(row.get("barcode") or ""),
                row.get("size") or "",
                int(row.get("qty") or 0),
                row.get("vendor_code") or "",
                row.get("nm_id") or "",
                row.get("name") or "",
                row.get("brand") or "",
                (row.get("status_label") or "") if first_of_box else "",
                (row.get("supply_ref") or "") if first_of_box else "",
            ]
        )
    _autosize(ws, STATE_HEADER)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_cells_xlsx(cells: list[dict[str, Any]]) -> bytes:
    """Сетка ячеек в формате A (тоже round-trip с `cells.parse_cells_file`)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ячейки"
    ws.append(CELLS_HEADER)
    _style_header(ws)
    for c in cells:
        ws.append(
            [
                c.get("warehouse_name") or c.get("warehouse") or "",
                c.get("cell_code") or c.get("code") or "",
                c.get("zone") or "",
                c.get("rack") or "",
                c.get("level") or "",
                c.get("pos") or "",
                "да" if c.get("is_active", True) else "нет",
                c.get("note") or "",
            ]
        )
    _autosize(ws, CELLS_HEADER)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_placement_xlsx(
    placements: list[dict[str, Any]],
    to_storage: list[dict[str, Any]],
    uncovered: list[dict[str, Any]],
    meta: dict[str, Any] | None = None,
) -> bytes:
    """«Лист размещения» для кладовщика (Фаза 2).

    Три листа: что куда поставить (по маршруту), что убрать на хранение, какие
    баркоды не попали в отбор.

    Каждый лист подписан складом и цифрами, а пустой лист ОБЯЗАН объяснить
    причину: файл с одной шапкой выглядит как сбой выгрузки, хотя чаще всего
    это штатное «свободных ячеек нет» или «весь ассортимент уже в отборе».
    """
    meta = meta or {}
    stats = meta.get("stats") or {}
    wh = meta.get("warehouse_name") or ""

    def head(ws: Any, title: str, note: str) -> None:
        ws.append([f"{title}{f' — склад «{wh}»' if wh else ''}"])
        ws["A1"].font = Font(bold=True, size=12)
        if note:
            ws.append([note])
        ws.append([])

    def table(ws: Any, header: list[str]) -> None:
        ws.append(header)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.freeze_panes = f"A{ws.max_row + 1}"

    def explain_empty(ws: Any, text: str) -> None:
        ws.append([])
        ws.append([text])

    cells_free = int(stats.get("cells_free") or 0)
    cells_total = int(stats.get("cells_total") or 0)
    covered = int(stats.get("barcodes_covered") or 0)
    total_bc = int(stats.get("barcodes_total") or 0)

    wb = Workbook()

    # ── Размещение ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Размещение"
    head(
        ws,
        "Лист размещения",
        f"ячеек {cells_total}, свободно {cells_free} · "
        f"в зоне отбора {covered} из {total_bc} баркодов",
    )
    header = ["Код ячейки", "Зона", "ШК короба", "Тип", "Баркоды", "Кол-во"]
    table(ws, header)
    for p in placements:
        step = int(p.get("step") or 0)
        ws.append(
            [
                p.get("cell_code") or "",
                p.get("zone") or "",
                p.get("box_code") or "",
                {1: "моно", 2: "сборный", 3: "пополнение"}.get(step, ""),
                (
                    f"{p.get('replenish_barcode')} (было {p.get('pick_qty_before')} шт)"
                    if step == 3
                    else ", ".join(str(b) for b in (p.get("covers") or []))
                ),
                int(p.get("total_qty") or 0),
            ]
        )
    if not placements:
        # Самая частая причина — 0 свободных ячеек. Пишем её словами.
        if cells_free == 0 and cells_total:
            why = (
                f"Размещать нечего: свободных ячеек нет — все {cells_total} заняты. "
                "Ячейки освободятся при отборе, тогда пополнение само подберёт "
                "короб на замену. Либо добавьте ячейки на «Карте склада»."
            )
        elif not cells_total:
            why = "Размещать нечего: на складе ещё не создано ни одной ячейки отбора."
        elif total_bc and covered >= total_bc:
            why = (
                "Размещать нечего: весь ассортимент уже доступен в зоне отбора, "
                "а пополнение отключено."
            )
        else:
            why = "Размещать нечего: нет коробов, которые можно переставить."
        explain_empty(ws, why)
    _autosize(ws, header)

    # ── На хранение ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet("На хранение")
    head(ws2, "На хранении (без адреса)", f"коробов {len(to_storage)}")
    header2 = ["ШК короба", "Бренд", "Кол-во", "Баркоды"]
    table(ws2, header2)
    for b in to_storage:
        ws2.append(
            [
                b.get("box_code") or "",
                b.get("brand") or "",
                int(b.get("total_qty") or 0),
                ", ".join(str(x) for x in (b.get("barcodes") or [])),
            ]
        )
    if not to_storage:
        explain_empty(ws2, "Все коробы разложены по ячейкам — на хранении ничего нет.")
    _autosize(ws2, header2)

    # ── Не покрыто ────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Не покрыто")
    head(ws3, "Не попали в зону отбора", f"баркодов {len(uncovered)}")
    header3 = ["Баркод", "nmID", "Артикул", "Название", "Кол-во на складе"]
    table(ws3, header3)
    for u in uncovered:
        ws3.append(
            [
                str(u.get("barcode") or ""),
                u.get("nm_id") or "",
                u.get("vendor_code") or "",
                u.get("name") or "",
                int(u.get("total_qty") or 0),
            ]
        )
    if not uncovered:
        explain_empty(
            ws3,
            "Пусто — это хорошо: весь ассортимент склада доступен в зоне отбора.",
        )
    _autosize(ws3, header3)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pick_xlsx(pick_orders: list[dict[str, Any]]) -> bytes:
    """Лист отбора: ОТДЕЛЬНЫЙ лист Excel на каждый кабинет.

    Решение пользователя — не сваливать кабинеты в один список: при упаковке и
    отгрузке их легко перепутать. Внутри листа строки идут по маршруту обхода
    склада, недостача — в конце и выделена текстом.
    """
    wb = Workbook()
    first = True
    header = [
        "№",
        "Ячейка",
        "ШК короба",
        "Баркод",
        "Артикул",
        "Название",
        "Взять, шт",
        "Отобрано",
        "Отметка",
    ]
    for order in pick_orders or []:
        title = str(order.get("cabinet_name") or "Отбор")[:31]
        ws = wb.active if first else wb.create_sheet(title)
        if first:
            ws.title = title
            first = False
        ws.append([f"Лист отбора — кабинет {order.get('cabinet_name')}"])
        ws.append(
            [
                f"заданий: {order.get('orders', 0)} · к отбору: {order.get('qty_required', 0)} шт"
                + (
                    f" · НЕДОСТАЧА: {order['shortage']} шт"
                    if order.get("shortage")
                    else ""
                )
            ]
        )
        ws.append([])
        ws.append(header)
        for cell in ws[4]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for i, line in enumerate(order.get("lines") or [], start=1):
            shortage = int(line.get("shortage") or 0)
            ws.append(
                [
                    i,
                    line.get("cell_code") or ("НЕТ НА СКЛАДЕ" if shortage else "хранение"),
                    line.get("box_code") or "",
                    str(line.get("barcode") or ""),
                    line.get("vendor_code") or "",
                    line.get("name") or "",
                    int(line.get("qty_required") or 0),
                    int(line.get("qty_picked") or 0),
                    "",
                ]
            )
        _autosize(ws, header)
        ws.freeze_panes = "A5"

    if first:  # ни одного листа не добавили
        ws = wb.active
        ws.title = "Отбор"
        ws.append(["Нечего отбирать"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
