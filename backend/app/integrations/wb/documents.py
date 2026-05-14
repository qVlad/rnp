"""WB Documents API integration.

Documents API даёт доступ к юридическим документам которые WB шлёт продавцу
отдельно от еженедельного отчёта реализации: уведомления о выкупе, УПД,
акты, претензии и др. Базовый host: `https://documents-api.wildberries.ru`.

Ключевые эндпоинты для налогового учёта:
  GET  /api/v1/documents/categories         — список доступных категорий
  GET  /api/v1/documents/list               — пагинированный список документов
  GET  /api/v1/documents/download           — скачать один (base64 ZIP)
  POST /api/v1/documents/download/all       — скачать пачку (1 req / 5 min!)

Этот модуль реализует только то, что нужно для tax_report: получение
`redeem-notification` (уведомлений о выкупе) с парсингом XLSX внутри ZIP.

Rate limits (для Base token, acc=3):
  list/categories/download    — 1 req / 10 sec, burst 5
  download/all                — 1 req / 5 min, burst 5 (используется в task)

Источник: WB API docs (dev.wildberries.ru), верифицировано 2026-05-14.
"""
from __future__ import annotations

import base64
import io
import re
import zipfile
from datetime import date, datetime
from decimal import Decimal
from typing import Any

from dateutil.parser import isoparse
from openpyxl import load_workbook

from app.integrations.wb.client import WbApiClient


# Канонические имена категорий WB Documents API
CATEGORY_REDEEM_NOTIFICATION = "redeem-notification"
CATEGORY_OFFSET_ACT = "actprofit"


async def list_documents(
    client: WbApiClient,
    category: str,
    date_from: date,
    date_to: date,
    *,
    limit: int = 50,
    offset: int = 0,
    locale: str = "ru",
) -> list[dict[str, Any]]:
    """`GET /api/v1/documents/list` — пагинированный список документов категории.

    Возвращает список dict-ов с полями:
      serviceName  — ID документа в WB ("redeem-notification-665788571")
      name         — категория ("redeem-notification")
      category     — человекочитаемое название ("Уведомление о выкупе")
      extensions   — ["zip"] обычно
      creationTime — ISO datetime когда WB создал документ
      viewed       — bool, открыт ли документ в кабинете
    """
    resp = await client.get(
        "/api/v1/documents/list",
        category="documents",
        params={
            "category": category,
            "beginTime": date_from.isoformat(),
            "endTime": date_to.isoformat(),
            "sort": "date",
            "order": "desc",
            "limit": limit,
            "offset": offset,
            "locale": locale,
        },
    )
    if not resp:
        return []
    return resp.get("data", {}).get("documents", []) or []


async def download_document(
    client: WbApiClient,
    service_name: str,
    *,
    extension: str = "zip",
) -> bytes:
    """`GET /api/v1/documents/download` — скачать один документ.

    Возвращает raw bytes ZIP-архива (после base64 decode). Лимит — 1/10 сек,
    подходит для on-demand download нескольких документов; для bulk используй
    `download_documents_batch` (1/5 мин но 50 за раз).
    """
    resp = await client.get(
        "/api/v1/documents/download",
        category="documents",
        params={"serviceName": service_name, "extension": extension},
    )
    if not resp:
        raise ValueError(f"empty response for {service_name}")
    b64 = resp.get("data", {}).get("document") or ""
    if not b64:
        raise ValueError(f"no document in response for {service_name}")
    return base64.b64decode(b64)


# ── Парсер XLSX уведомления о выкупе ────────────────────────────────────
# Формат WB (стабильный 2024-2026):
#   A3: «УВЕДОМЛЕНИЕ О ВЫКУПЕ №<номер> от <дата>»
#   Строка 10: заголовки (№ п/п, Артикул, Наименование, Кол-во, Сумма, ...)
#   Строки 11..N: товарные позиции
#   Строка с «Итого:» в A: финальный total, E = сумма выкупа (вкл. НДС)

_HEADER_RE = re.compile(
    r"УВЕДОМЛЕНИЕ О ВЫКУПЕ\s*№\s*(?P<num>\d+).*?от\s+(?P<date>\d{4}-\d{2}-\d{2}|\d{2}\.\d{2}\.\d{4})",
    re.IGNORECASE | re.DOTALL,
)


def _parse_ru_decimal(value: Any) -> Decimal | None:
    """Парсит русский числовой формат: "16 064,07" → Decimal("16064.07").
    Поддерживает обычный пробел, NBSP (\xa0), запятую как десятичный
    разделитель. Если на входе уже число — конвертирует напрямую."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except Exception:
            return None
    s = str(value).strip()
    # Убираем все варианты пробелов-разделителей тысяч
    s = s.replace("\xa0", "").replace(" ", "").replace(" ", "")
    # Запятая → точка (русский десятичный)
    s = s.replace(",", ".")
    try:
        return Decimal(s)
    except Exception:
        return None


def _parse_date_loose(s: str) -> date | None:
    """Поддерживаем оба формата которые WB использует: 2026-05-14 и 14.05.2026."""
    s = (s or "").strip()
    if not s:
        return None
    try:
        if "." in s and len(s.split(".")[0]) <= 2:
            d, m, y = s.split(".")
            return date(int(y), int(m), int(d))
        return isoparse(s).date()
    except (ValueError, IndexError):
        return None


def parse_redeem_notification(zip_bytes: bytes) -> dict[str, Any]:
    """Распарсить ZIP с уведомлением о выкупе. Возвращает:
      {
        "notification_number": "713648439",
        "notification_date": date(2026, 5, 4),
        "total_sum_with_vat": Decimal("16064.07"),
        "items": [{"vendor_code", "name", "qty", "sum", "vat_rate", "kiz"}, ...],
        "service_name": "redeem-notification-713648439",
      }

    Бросает ValueError если структура XLSX не распознана.
    """
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as z:
        xlsx_name = next((n for n in z.namelist() if n.endswith(".xlsx")), None)
        if xlsx_name is None:
            raise ValueError("no .xlsx inside ZIP")
        xlsx_bytes = z.read(xlsx_name)

    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    s = wb.active

    # 1) Шапка с номером и датой
    header_text = ""
    for r in range(1, min(10, s.max_row) + 1):
        v = s.cell(row=r, column=1).value
        if v and "УВЕДОМЛЕНИЕ О ВЫКУПЕ" in str(v).upper():
            header_text = str(v)
            break
    m = _HEADER_RE.search(header_text or "")
    if not m:
        raise ValueError(f"can't parse header: {header_text!r}")
    notification_number = m.group("num")
    notification_date = _parse_date_loose(m.group("date"))
    if notification_date is None:
        raise ValueError(f"can't parse date from header: {header_text!r}")

    # 2) Найти строку с «Итого:» и индекс заголовков
    total_row = None
    header_row = None
    for r in range(1, s.max_row + 1):
        cell_a = (s.cell(row=r, column=1).value or "")
        cell_a_str = str(cell_a).strip().lower()
        if cell_a_str.startswith("итого"):
            total_row = r
            break
        if cell_a_str == "№ \nп/п" or "артикул" in str(s.cell(row=r, column=2).value or "").lower():
            header_row = r
    if total_row is None:
        raise ValueError("no Итого row found")

    # E (col 5) = «Сумма выкупа, руб., (вкл. НДС)»
    # WB иногда хранит сумму как форматированную строку "16 064,07" (русский
    # формат: NBSP/обычный пробел = разделитель тысяч, запятая = десятичный),
    # иногда как число. Нормализуем оба случая.
    total_value = s.cell(row=total_row, column=5).value
    total_sum = _parse_ru_decimal(total_value)
    if total_sum is None:
        raise ValueError(f"can't parse total sum: {total_value!r}")

    # 3) Товарные позиции (между header_row + 1 и total_row - 1)
    items: list[dict[str, Any]] = []
    start_row = (header_row + 1) if header_row else 11
    for r in range(start_row, total_row):
        idx = s.cell(row=r, column=1).value
        if idx is None:
            continue
        vendor_code = s.cell(row=r, column=2).value
        name = s.cell(row=r, column=3).value
        qty = s.cell(row=r, column=4).value
        sum_val = s.cell(row=r, column=5).value
        vat_rate = s.cell(row=r, column=6).value
        vat_sum = s.cell(row=r, column=7).value
        kiz = s.cell(row=r, column=8).value
        try:
            qty_int = int(qty) if qty is not None else 0
        except (ValueError, TypeError):
            qty_int = 0
        try:
            sum_dec = float(sum_val) if sum_val is not None else 0.0
        except (ValueError, TypeError):
            sum_dec = 0.0
        items.append({
            "vendor_code": str(vendor_code or "").strip(),
            "name": str(name or "").strip(),
            "qty": qty_int,
            "sum": sum_dec,
            "vat_rate": str(vat_rate or "").strip(),
            "vat_sum": str(vat_sum or "").strip() if vat_sum not in (None, "—") else None,
            "kiz": str(kiz or "").strip() if kiz not in (None, "—") else None,
        })

    return {
        "notification_number": notification_number,
        "notification_date": notification_date,
        "total_sum_with_vat": total_sum,
        "items": items,
        "service_name": f"redeem-notification-{notification_number}",
    }


# ── Парсер XLSX акта взаимозачёта (actprofit) ───────────────────────────
# Generic-парсер: ищет шапку с номером + датой, и строку «Итого» с суммой.
# Точная структура неизвестна (у клиента нет таких документов в живом
# tenants за последние 5 месяцев) — берём согласованный подход с redeem.

_OFFSET_HEADER_RE = re.compile(
    r"(?:АКТ\s+ВЗАИМОЗАЧ[ЕЁ]ТА|АКТ\s+ЗАЧ[ЕЁ]ТА)\s*№?\s*(?P<num>\d+).*?от\s+(?P<date>\d{4}-\d{2}-\d{2}|\d{2}\.\d{2}\.\d{4})",
    re.IGNORECASE | re.DOTALL,
)


def parse_offset_act(zip_bytes: bytes) -> dict[str, Any]:
    """Распарсить ZIP с актом взаимозачёта. Возвращает:
      {
        "act_number": "...",
        "act_date": date(...),
        "total_sum": Decimal(...),
        "items": [...],
        "service_name": "actprofit-...",
      }

    Парсер написан как fallback с поддержкой нескольких форматов:
    1. Шапка с «Акт взаимозачёта/зачёта №<n> от <date>»
    2. Сумма в строке «Итого:» (любая колонка с числом)

    При появлении первого реального документа на проде — подкорректировать
    логику под фактический формат WB."""
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as z:
        xlsx_name = next((n for n in z.namelist() if n.endswith(".xlsx")), None)
        if xlsx_name is None:
            raise ValueError("no .xlsx inside ZIP")
        xlsx_bytes = z.read(xlsx_name)

    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    s = wb.active

    # 1) Шапка
    header_text = ""
    for r in range(1, min(15, s.max_row) + 1):
        v = s.cell(row=r, column=1).value
        if v and ("АКТ" in str(v).upper() and "ЗАЧ" in str(v).upper()):
            header_text = str(v)
            break
    m = _OFFSET_HEADER_RE.search(header_text or "")
    if m:
        act_number = m.group("num")
        act_date = _parse_date_loose(m.group("date"))
    else:
        # Fallback — взять номер из serviceName и текущую дату
        act_number = "unknown"
        act_date = None

    if act_date is None:
        raise ValueError(f"can't parse date from header: {header_text!r}")

    # 2) Сумма из «Итого:»
    total_sum: Decimal | None = None
    for r in range(s.max_row, 0, -1):
        cell_a = str(s.cell(row=r, column=1).value or "").strip().lower()
        if cell_a.startswith("итого"):
            # Просканировать строку и взять самое большое число (вероятный total)
            for col in range(2, s.max_column + 1):
                v = s.cell(row=r, column=col).value
                parsed = _parse_ru_decimal(v)
                if parsed is not None and (total_sum is None or parsed > total_sum):
                    total_sum = parsed
            if total_sum is not None:
                break
    if total_sum is None:
        raise ValueError("no Итого row with parseable sum found")

    # 3) Items — собираем все непустые строки между шапкой и «Итого»
    items: list[dict[str, Any]] = []
    for r in range(2, s.max_row + 1):
        row_values = [s.cell(row=r, column=c).value for c in range(1, min(s.max_column, 10) + 1)]
        if all(v in (None, "") for v in row_values):
            continue
        cell_a = str(row_values[0] or "").strip().lower()
        if cell_a.startswith("итого"):
            break
        # Пропускаем шапку
        if any(kw in str(row_values).upper() for kw in ("АКТ", "СУММА", "ОБЩЕСТВО", "ИНН")):
            continue
        items.append({f"col_{i+1}": str(v) if v is not None else None for i, v in enumerate(row_values)})

    return {
        "act_number": act_number,
        "act_date": act_date,
        "total_sum": total_sum,
        "items": items,
        "service_name": f"actprofit-{act_number}",
    }
